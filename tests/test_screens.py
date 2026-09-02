"""Does the app actually render?

`scripts/check_db.py` proves the data is sound. It cannot tell you that a
template throws, that a query returns the wrong rows, or that a route 500s.
This file is that half of the check.

Add a test here for every screen as it gets built — see docs/BACKLOG.md, which
lists the acceptance criteria each one has to meet.
"""

from __future__ import annotations

import re
from pathlib import Path

import pytest
from conftest import STAFF_NAME, STAFF_PIN, days

ROOT = Path(__file__).resolve().parent.parent


# --------------------------------------------------------------- signing in

def test_login_lists_active_staff(anon_client, staff):
    body = anon_client.get("/login").text
    assert STAFF_NAME in body


def test_login_hides_inactive_staff(anon_client, staff, db):
    db.execute("UPDATE users SET active = 0 WHERE id = ?", (staff["id"],))
    db.commit()
    assert STAFF_NAME not in anon_client.get("/login").text


def test_login_shows_a_keypad(anon_client, staff):
    """Big targets, no keyboard needed — one-handed on an iPad."""
    body = anon_client.get(f"/login?user={staff['id']}").text
    for digit in "0123456789":
        assert f'data-digit="{digit}"' in body


def test_correct_pin_signs_in_and_lands_on_due(anon_client, staff):
    response = anon_client.post(
        "/login", data={"user_id": staff["id"], "pin": STAFF_PIN}, follow_redirects=False
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/"
    assert anon_client.get("/").status_code == 200


def test_wrong_pin_says_so_plainly(anon_client, staff):
    response = anon_client.post(
        "/login", data={"user_id": staff["id"], "pin": "9999"}, follow_redirects=False
    )
    assert response.status_code == 200            # re-rendered, not redirected
    assert "did not match" in response.text
    # Must not leak which half was wrong, and must never echo the PIN back.
    assert "9999" not in response.text
    assert anon_client.get("/", follow_redirects=False).status_code == 303


def test_wrong_pin_does_not_lock_anybody_out(anon_client, staff):
    """PINs are accountability, not security. No lockout, by decision."""
    for _ in range(5):
        anon_client.post("/login", data={"user_id": staff["id"], "pin": "0000"})
    response = anon_client.post(
        "/login", data={"user_id": staff["id"], "pin": STAFF_PIN}, follow_redirects=False
    )
    assert response.status_code == 303


def test_no_session_redirects_to_login(anon_client):
    for path in ("/", "/scan", "/products", "/sheet", "/settings"):
        response = anon_client.get(path, follow_redirects=False)
        assert response.status_code == 303, path
        assert response.headers["location"] == "/login", path


def test_login_and_static_are_reachable_signed_out(anon_client):
    assert anon_client.get("/login").status_code == 200
    assert anon_client.get("/static/css/app.css").status_code == 200


def test_signed_in_name_is_on_every_page(client):
    """Proof the user reached the route: writes stamp this id, so it has to be
    there without the route re-reading the cookie."""
    assert STAFF_NAME in client.get("/").text


def test_logout_clears_the_session(client):
    response = client.get("/logout", follow_redirects=False)
    assert response.status_code == 303
    assert client.get("/", follow_redirects=False).status_code == 303


def test_login_page_works_with_no_staff_at_all(anon_client):
    """A fresh database has nobody in it. The page must still explain itself."""
    response = anon_client.get("/login")
    assert response.status_code == 200
    assert "add_user.py" in response.text


# ------------------------------------------------------------------- home

def test_home_renders_when_empty(client):
    """A fresh database is a valid state. The home screen must not need data.

    It also must not say "nothing due in the next 7 days" here. That sentence
    is reassurance — all clear, nothing coming up — and on a database with no
    products in it the truth is the opposite: the app has never been used. This
    is the first screen a new shop sees, so it says so and points at the one
    thing to do next.
    """
    response = client.get("/")
    assert response.status_code == 200
    assert "BP Tecoma" in response.text
    assert "Nothing is being tracked yet" in response.text
    assert 'href="/scan"' in response.text
    assert "Nothing due in the next" not in response.text


def test_a_stocked_shop_with_a_quiet_week_still_says_nothing_is_due(client, sample, db):
    """The other half of the pair. Products exist and none is due soon, which is
    a real all-clear and must keep reading as one."""
    db.execute("DELETE FROM batches")
    db.commit()
    response = client.get("/")
    assert "Nothing due in the next 7 days" in response.text
    assert "Nothing is being tracked yet" not in response.text


def test_home_shows_due_items(client, sample):
    response = client.get("/")
    assert response.status_code == 200
    assert "C/RIDGE WATER 1L" in response.text          # due in 2 days
    assert "Monster Ultra Zero 500ml" in response.text  # 4 days overdue


def test_home_puts_items_beyond_the_window_under_their_own_heading(client, sample, db):
    """Changed 16 Aug with the day bands, and it is a reversal worth stating.

    The 30-day batch used to be absent from this screen. It is now on it, under
    "More than a week" — the window stopped deciding what is visible and started
    deciding where the per-day bands stop. Only its product's other batch puts
    that name on the page, so check the date instead.
    """
    far_off = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["outside"],)
    ).fetchone()[0]
    body = client.get("/").text
    assert _au(far_off) in body
    assert "More than a week" in body
    # ...and below every band that is somebody's job sooner.
    assert body.index("Past date") < body.index("More than a week")
    assert body.index("Tomorrow") < body.index("More than a week")


def test_home_shows_a_discounted_batch(client, sample, db):
    """Inverted 13 Aug, and this is the point of the status change.

    This used to assert the resolved batch was hidden, because 'pulled' and
    'sold' meant gone-but-recorded. Both are removed: a discounted batch is
    still on the shelf, still needs watching, and is the only resolution left.
    A batch that is genuinely finished with is deleted, and a deleted row cannot
    be hidden because it does not exist.
    """
    discounted = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["discounted"],)
    ).fetchone()[0]
    assert _au(discounted) in client.get("/").text


def test_home_includes_the_window_edge(client, sample, db):
    """Exactly 7 days out is inside the window, not outside it. Off-by-one here
    silently hides a day's worth of stock."""
    edge = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["due_edge"],)
    ).fetchone()[0]
    assert _au(edge) in client.get("/").text


def test_home_puts_overdue_first(client, sample):
    """Past-date items sit above upcoming ones. Normal state, top of the page."""
    body = client.get("/").text
    assert body.index("Past date") < body.index("Tomorrow")


def test_overdue_is_not_an_error(client, sample):
    """583 batches imported already expired. Overdue is expected, not a warning."""
    body = client.get("/").text
    for alarm in ("Error", "error", "Warning", "warning"):
        assert alarm not in body


def test_uncategorised_product_renders_blank(client, sample):
    """category_id IS NULL is normal. No 'Uncategorised' label anywhere."""
    body = client.get("/").text
    assert "Uncategorised" not in body
    assert "None" not in body  # a NULL leaking into the template


def test_home_counts_only_live_batches(client, sample):
    """3 products; all 5 batches are live now that discounted is the only
    resolution — there is no status that means present-but-not-counted."""
    body = client.get("/").text
    assert "3 products, 5 being tracked" in body


def test_every_due_row_has_buttons_for_both_actions(client, sample):
    """Buttons, not gestures — swipe was removed on 13 Aug after the iPad test.

    A gesture whose threshold a person cannot see is a guess, and the two guesses
    were "delete" and "discount". The buttons say what they do.
    """
    body = client.get("/").text
    # All five live rows are on the page from 16 Aug: overdue, the discounted +1,
    # +2, +7, and the +30 under "More than a week". Being further off is not a
    # reason to be unable to act on it.
    assert body.count("/delete") == 5                   # every row shown
    assert body.count('value="discounted"') == 4        # not the already-discounted one
    assert body.count("/full-price") == 1               # only the discounted one
    assert 'value="/"' in body                          # comes back to the Due screen


def test_the_due_screen_needs_no_javascript(client, sample):
    """It is the busiest read-only screen in the shop, and now carries no script.

    The delete confirmation used to be built by swipe.js. It is a <details> the
    server chooses to render, so the rule lives in needs_confirmation() and there
    is no script that could disagree with it — or fail to load and take the
    confirmation with it.
    """
    body = client.get("/").text
    assert "<script" not in body
    assert "swipe.js" not in body
    assert client.get("/static/js/swipe.js").status_code == 404


def test_the_delete_confirmation_is_only_on_rows_that_need_asking(client, sample, db):
    """Live batches in `sample`: overdue (past, active), due_soon (+2, active),
    due_edge (+7, active), outside (+30, active, now shown under "More than a
    week"), discounted (+1). All five are on the page from 16 Aug, and the three
    active in-date ones ask before deleting — the past-date one does not, and
    neither does the discounted one. The rule is about the batch, not about
    which band it happens to sit in.
    """
    body = client.get("/").text
    assert body.count('<details class="confirm">') == 3
    collapsed = " ".join(body.split())
    assert "This expires in 2 days. Delete it?" in collapsed
    assert "This expires in 7 days. Delete it?" in collapsed
    assert "This expires in 30 days. Delete it?" in collapsed


@pytest.mark.parametrize("path,expected_class", [
    ("/", 'class="item-row"'),                       # Due: shares the row with buttons
    ("/products", 'class="item-row item-row-full"'),  # Products: no buttons, whole row
])
def test_tapping_a_row_covers_the_thumbnail(client, sample, path, expected_class):
    """The picture is the first thing a thumb reaches, and it did nothing before.

    The link used to wrap only the words, leaving the thumbnail outside it. Fixed
    on the Due screen on 13 Aug and on the products list straight after, because
    fixing one and not the other is exactly the kind of thing a parametrised test
    would have caught the first time.
    """
    body = client.get(path).text
    assert expected_class in body, path

    # Sliced from the opening <a> rather than from the <li>: everything between
    # the two contains the thumbnail whether or not the link wraps it, so the
    # wider slice passes against the bug it is meant to catch.
    link = body.split('<a class="item-row')[1].split("</a>")[0]
    assert "thumb" in link, f"{path}: the thumbnail is outside the link again"

    css = client.get("/static/css/app.css").text
    assert ".item-row {" in css


def test_the_products_row_link_reaches_the_date_column_too(client, sample):
    """Nothing on this row needs to stay outside the link, so nothing does."""
    body = client.get("/products").text
    link = body.split('class="item-row item-row-full"')[1].split("</a>")[0]
    assert "item-body" in link
    assert "item-right" in link, "the date column is outside the tap target"
    assert ".item-row-full" in client.get("/static/css/app.css").text


def test_photo_placeholder_when_no_image(client, sample):
    """Lists must not reflow when photos are backfilled months later."""
    assert "thumb-empty" in client.get("/").text


def test_dates_are_australian_on_the_page(client, sample):
    """Never US format. 4 Sep, not Sep 4 or 09/04."""
    body = client.get("/").text
    assert _au(days(2)) in body


def test_home_filters_by_category(client, sample):
    body = client.get(f"/?category={sample['cat_id']}").text
    assert "Monster Ultra Zero 500ml" in body
    assert "C/RIDGE WATER 1L" not in body


def test_home_filters_to_the_uncategorised(client, sample):
    """A normal state with a filter of its own — and still no 'Uncategorised'
    row anywhere, because NULL is how it is stored."""
    body = client.get("/?category=none").text
    assert "C/RIDGE WATER 1L" in body
    assert "Monster Ultra Zero 500ml" not in body
    assert "Uncategorised" not in body


def test_home_shows_no_filter_bar_until_a_category_exists(client, db):
    """The list starts empty and grows by itself. No empty chrome before then."""
    assert 'class="filters"' not in client.get("/").text


def test_home_rows_link_to_the_product(client, sample):
    assert f'href="/products/{sample["products"]["monster"]}"' in client.get("/").text


# ------------------------------------------------------------ scan and add

def test_scan_opens_with_the_cursor_in_the_barcode_field(client):
    body = client.get("/scan").text
    assert 'id="barcode"' in body
    assert "autofocus" in body.split('id="barcode"')[1].split(">")[0]


def test_known_barcode_shows_the_product_and_focuses_the_date(client, sample):
    body = client.get("/scan?barcode=9300601234567").text
    assert "Monster Ultra Zero 500ml" in body
    assert 'id="expiry"' in body
    assert "autofocus" in body.split('id="expiry"')[1].split(">")[0]
    assert 'id="name"' not in body            # nothing to re-type for a known product


def test_unknown_barcode_asks_for_a_name(client, sample):
    body = client.get("/scan?barcode=9999999999999").text
    assert 'id="name"' in body
    assert 'id="expiry"' in body


def test_a_typed_word_is_refused_and_creates_nothing(client, db):
    """The punch-list bug itself. Products are never deleted, so junk is forever."""
    before = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    response = client.post(
        "/scan/add",
        data={"barcode": "cool ridge water", "name": "Cool Ridge Water 600ml",
              "expiry_date": days(10)},
        follow_redirects=False,
    )
    assert response.status_code == 200            # re-rendered with the reason
    assert "digits only" in response.text
    assert db.execute("SELECT COUNT(*) FROM products").fetchone()[0] == before
    assert db.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == 0


def test_an_invalid_barcode_never_opens_the_new_product_form(client):
    """Nobody should type a name for something that will be refused on save."""
    body = client.get("/scan?barcode=cool ridge water").text
    assert "digits only" in body
    assert 'id="name"' not in body
    assert 'id="barcode"' in body                 # back to step one, ready to rescan


def test_a_barcode_of_the_wrong_length_says_the_length(client, db):
    response = client.post(
        "/scan/add", data={"barcode": "12345", "name": "Too Short", "expiry_date": days(10)}
    )
    assert "6 to 18 digits" in response.text
    assert "That one is 5" in response.text
    assert db.execute("SELECT COUNT(*) FROM products").fetchone()[0] == 0


def test_a_gun_prefixed_new_barcode_is_stored_without_the_prefix(client, db, staff):
    """The stored barcode is the code, not the gun's announcement of it."""
    client.post(
        "/scan/add",
        data={"barcode": "]C10118721274620198", "name": "Golden Gay Time Lamington",
              "expiry_date": days(12)},
    )
    row = db.execute(
        "SELECT barcode FROM products WHERE name = 'Golden Gay Time Lamington'"
    ).fetchone()
    assert row is not None, "the add was refused"
    assert row["barcode"] == "0118721274620198"


def test_adding_writes_a_batch_stamped_with_the_signed_in_user(client, sample, db, staff):
    response = client.post(
        "/scan/add",
        data={"barcode": "9300601234567", "expiry_date": days(45)},
        follow_redirects=False,
    )
    assert response.status_code == 303
    row = db.execute(
        "SELECT b.added_by, b.status, b.quantity FROM batches b "
        "WHERE b.product_id = ? AND b.expiry_date = ?",
        (sample["products"]["monster"], days(45)),
    ).fetchone()
    assert row is not None
    assert row["added_by"] == staff["id"]
    assert row["status"] == "active"
    assert row["quantity"] == 1


def test_adding_an_unknown_barcode_creates_the_product_too(client, db, staff):
    client.post(
        "/scan/add",
        data={"barcode": "9300600000123", "name": "Solo Energy Lemon 500ml",
              "expiry_date": days(20)},
    )
    product = db.execute(
        "SELECT id, name, category_id FROM products WHERE barcode = ?",
        ("9300600000123",),
    ).fetchone()
    assert product is not None
    assert product["name"] == "Solo Energy Lemon 500ml"
    assert product["category_id"] is None      # no category is a normal state

    # The person is recorded on the batch, not on the product — dropped 13 Aug.
    assert db.execute(
        "SELECT added_by FROM batches WHERE product_id = ?", (product["id"],)
    ).fetchone()["added_by"] == staff["id"]
    assert db.execute(
        "SELECT COUNT(*) FROM batches WHERE product_id = ?", (product["id"],)
    ).fetchone()[0] == 1


def test_duplicate_is_caught_before_insert_with_the_date(client, sample, db):
    """The core rule. Not a database error, and no offer to raise a quantity."""
    existing = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["overdue"],)
    ).fetchone()[0]

    response = client.post(
        "/scan/add", data={"barcode": "9300601234567", "expiry_date": existing}
    )
    assert response.status_code == 200
    assert f"Already tracked — expires {_au(existing)}" in response.text
    assert "quantity" not in response.text.lower()
    assert "IntegrityError" not in response.text

    assert db.execute(
        "SELECT COUNT(*) FROM batches WHERE product_id = ? AND expiry_date = ?",
        (sample["products"]["monster"], existing),
    ).fetchone()[0] == 1


def test_a_duplicate_still_keeps_the_category_that_was_typed(client, sample, db):
    """Proof the app checks before inserting rather than letting the index
    raise: an IntegrityError would roll the category back with it."""
    duplicate_date = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["due_soon"],)
    ).fetchone()[0]

    response = client.post(
        "/scan/add",
        data={"barcode": "9300609876543", "category": "Water", "expiry_date": duplicate_date},
    )
    assert "Already tracked" in response.text
    assert db.execute(
        """SELECT c.name FROM products p JOIN categories c ON c.id = p.category_id
            WHERE p.barcode = ?""",
        ("9300609876543",),
    ).fetchone()[0] == "Water"


def test_a_discounted_batch_still_blocks_its_date(client, sample, db):
    """Rewritten 13 Aug. A discounted batch is live, so the date is taken.

    The old version used the 'pulled' batch to prove a resolved date came free
    again. Nothing is resolved-but-present any more, so the only live status
    besides active is 'discounted' — and that item is on the shelf, so recording
    it twice is the duplicate this app exists to prevent.
    """
    taken = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["discounted"],)
    ).fetchone()[0]

    response = client.post(
        "/scan/add",
        data={"barcode": "9300601111111", "expiry_date": taken},
        follow_redirects=False,
    )
    assert response.status_code == 200            # refused, re-rendered
    assert "Already tracked" in response.text


def test_deleting_a_batch_frees_its_date(client, sample, db):
    """Deletion is what frees a date now, and it frees it immediately.

    This is why idx_batches_unique_live can stay partial and still be correct:
    with no resolved-but-present rows, the deleted batch is simply not there.
    """
    batch_id = sample["batches"]["discounted"]
    freed = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (batch_id,)
    ).fetchone()[0]

    client.post(f"/products/{sample['products']['curly']}/batches/{batch_id}/delete")

    response = client.post(
        "/scan/add",
        data={"barcode": "9300601111111", "expiry_date": freed},
        follow_redirects=False,
    )
    assert response.status_code == 303, "the freed date was still blocked"
    assert db.execute(
        "SELECT COUNT(*) FROM batches WHERE product_id = ? AND expiry_date = ? "
        "AND status = 'active'",
        (sample["products"]["curly"], freed),
    ).fetchone()[0] == 1


def test_after_saving_the_form_is_back_at_the_barcode_field(client, sample):
    response = client.post(
        "/scan/add",
        data={"barcode": "9300601234567", "expiry_date": days(46)},
        follow_redirects=True,
    )
    body = response.text
    assert "Saved" in body
    assert 'id="barcode"' in body
    assert 'id="expiry"' not in body          # nothing half-finished left on screen


def test_a_rejected_date_does_not_lose_the_typed_name(client):
    body = client.post(
        "/scan/add",
        data={"barcode": "9300600000999", "name": "Chobani Greek Yoghurt 170g",
              "expiry_date": ""},
    ).text
    assert "Chobani Greek Yoghurt 170g" in body
    assert "Enter the expiry date." in body


def test_a_mistyped_year_is_refused(client, sample):
    body = client.post(
        "/scan/add", data={"barcode": "9300601234567", "expiry_date": "0226-09-14"}
    ).text
    assert "Check the year" in body


def test_scan_never_renders_a_us_date(client, sample, db):
    existing = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["overdue"],)
    ).fetchone()[0]
    body = client.post(
        "/scan/add", data={"barcode": "9300601234567", "expiry_date": existing}
    ).text
    assert "/" not in body.split("Already tracked")[1].split("<")[0]


# ------------------------------------------------------- inline categories

def test_category_input_suggests_the_existing_ones(client, sample):
    body = client.get("/scan?barcode=9300609876543").text
    assert '<datalist id="category-list">' in body
    assert '<option value="Energy Drinks">' in body


def test_typing_a_new_category_creates_it_against_the_product(client, db, staff):
    client.post(
        "/scan/add",
        data={"barcode": "9300600000456", "name": "Bulla Ice Cream 2L",
              "category": "Frozen", "expiry_date": days(30)},
    )
    category = db.execute(
        "SELECT id, created_by FROM categories WHERE name = 'Frozen'"
    ).fetchone()
    assert category is not None
    assert category["created_by"] == staff["id"]
    assert db.execute(
        "SELECT category_id FROM products WHERE barcode = ?", ("9300600000456",)
    ).fetchone()[0] == category["id"]


def test_a_case_variant_picks_the_existing_category(client, sample, db):
    """Typing 'energy drinks' must find 'Energy Drinks', not collide with it."""
    client.post(
        "/scan/add",
        data={"barcode": "9300609876543", "category": "energy drinks",
              "expiry_date": days(31)},
    )
    assert db.execute("SELECT COUNT(*) FROM categories").fetchone()[0] == 1
    assert db.execute(
        "SELECT category_id FROM products WHERE barcode = ?", ("9300609876543",)
    ).fetchone()[0] == sample["cat_id"]


def test_a_category_covers_batches_recorded_earlier(client, sample, db):
    """It attaches to the barcode, so it reaches back over old batches."""
    client.post(
        "/scan/add",
        data={"barcode": "9300601111111", "category": "Biscuits", "expiry_date": days(32)},
    )
    rows = db.execute(
        """SELECT c.name FROM batches b
             JOIN products p ON p.id = b.product_id
             JOIN categories c ON c.id = p.category_id
            WHERE b.product_id = ?""",
        (sample["products"]["curly"],),
    ).fetchall()
    assert len(rows) == 3                       # every batch of that barcode, old ones too
    assert all(r["name"] == "Biscuits" for r in rows)


def test_a_blank_category_is_never_complained_about(client, sample, db):
    response = client.post(
        "/scan/add",
        data={"barcode": "9300609876543", "category": "  ", "expiry_date": days(33)},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert db.execute(
        "SELECT category_id FROM products WHERE barcode = ?", ("9300609876543",)
    ).fetchone()[0] is None


def test_no_uncategorised_option_is_offered(client, sample):
    assert "Uncategorised" not in client.get("/scan?barcode=9300609876543").text


# ------------------------------------------------- product list and detail

def test_product_list_shows_everything_by_default(client, sample):
    body = client.get("/products").text
    assert "Monster Ultra Zero 500ml" in body
    assert "C/RIDGE WATER 1L" in body
    assert "Arnott" in body


def test_search_is_case_insensitive(client, sample):
    assert "Monster Ultra Zero 500ml" in client.get("/products?q=MONSTER").text
    assert "Monster Ultra Zero 500ml" in client.get("/products?q=monster").text


def test_search_tolerates_surrounding_whitespace(client, sample):
    """Stored names have trailing spaces; typed queries have stray ones."""
    body = client.get("/products?q=%20%20water%20%20").text
    assert "C/RIDGE WATER 1L" in body


def test_search_finds_the_abbreviated_name(client, sample):
    """'cool ridge' cannot substring-match 'C/RIDGE WATER 1L' — the shop's
    abbreviation is not derivable — so the word that does match carries it."""
    body = client.get("/products?q=cool+ridge").text
    assert "C/RIDGE WATER 1L" in body


def test_search_handles_a_curly_apostrophe_either_way(client, sample):
    """The export has Arnott’s with a curly one. Nobody types that."""
    assert "Tim Tam" in client.get("/products?q=arnott%27s").text
    assert "Tim Tam" in client.get("/products?q=arnott%E2%80%99s").text


def test_search_by_barcode(client, sample):
    body = client.get("/products?q=9300609876543").text
    assert "C/RIDGE WATER 1L" in body
    assert "Monster Ultra Zero 500ml" not in body


def test_search_that_matches_nothing_says_so(client, sample):
    assert "Nothing matches that." in client.get("/products?q=zzzzzz").text


def test_an_empty_catalogue_is_not_a_failed_search(client, staff):
    """No `sample` fixture here — this is a shop that has not started.

    "Nothing matches that." answers a question nobody asked. Products are
    created by scanning, so the empty screen says that and offers the way in
    rather than implying a search came back empty.
    """
    body = client.get("/products").text
    assert "No products yet" in body
    assert 'href="/scan"' in body
    assert "Nothing matches that." not in body


def test_a_search_that_finds_nothing_still_says_so(client, sample):
    """The other half: with products present, a miss is a miss."""
    body = client.get("/products?q=zzzzzz").text
    assert "Nothing matches that." in body
    assert "No products yet" not in body


def test_search_treats_a_percent_sign_as_a_character(client, sample, db):
    """`%` is LIKE's "anything" wildcard, so it used to match every product in
    the shop. The real export has "70% dark chocolate" and "99% sugar free" in
    it, and someone typing 70% wants those, not all 945 rows."""
    db.execute(
        "INSERT INTO products (barcode, name) VALUES (?, ?)",
        ("9300600000070", "Tony’s Chocolonely 70% dark chocolate"),
    )
    db.commit()

    # The wildcard alone finds the one product with a percent sign in its name,
    # not the whole catalogue.
    body = client.get("/products?q=%25").text
    assert "Chocolonely" in body
    assert "Monster Ultra Zero 500ml" not in body

    # And the way a person would actually type it.
    assert "Chocolonely" in client.get("/products?q=70%25").text


def test_search_treats_an_underscore_as_a_character(client, sample):
    """`_` is LIKE's "any one character". No product name here has one, so the
    honest answer is nothing — not everything."""
    assert "Nothing matches that." in client.get("/products?q=_").text


def test_live_search_returns_only_the_rows_fragment(client, sample):
    """search-live.js fetches ?partial=1 and drops the result straight into the
    page as someone types, so the fragment has to be the rows alone — no header,
    and not the search form again (which would nest one form inside another)."""
    body = client.get("/products?q=cool+ridge&partial=1").text
    assert "C/RIDGE WATER 1L" in body            # the scored match still surfaces
    assert "<html" not in body                   # no base chrome
    assert 'id="product-search"' not in body     # not the search form
    assert 'id="product-results"' not in body    # not the container it fills


def test_live_search_fragment_runs_the_same_search(client, sample):
    """The partial is the ordinary query, so it agrees with the full page —
    a miss says so, and a filter that excludes the water really excludes it."""
    assert "Nothing matches that." in client.get("/products?q=zzzzzz&partial=1").text
    assert "C/RIDGE WATER 1L" not in client.get("/products?q=monster&partial=1").text


def test_the_products_page_enhances_search_and_scroll(client, sample):
    """The full page carries the two enhancement scripts and the container they
    act on. Without them the plain form still submits; with them the list
    filters as you type and returning to it keeps your place."""
    body = client.get("/products").text
    assert "search-live.js" in body
    assert "list-scroll.js" in body
    assert 'id="product-results"' in body


def test_pages_are_compressed_on_the_way_out(client, sample):
    """The Due screen is 485 KB of HTML on the shop's data and ~20 KB gzipped —
    about 2.5 seconds of a weak WiFi link, on the first page of every session.
    Losing this middleware would be invisible on the dev laptop and painful in
    the shop, which is exactly the kind of regression worth a test."""
    response = client.get("/products?all=1", headers={"Accept-Encoding": "gzip"})
    assert response.headers.get("content-encoding") == "gzip"


def test_a_versioned_asset_can_be_kept_but_an_unversioned_one_cannot(client, sample):
    """asset_url() and photo_url() stamp the file's mtime into the URL, so those
    bytes can never change and a device may keep them for a year. The icons are
    referenced by plain path, so they must not be frozen — a redrawn icon has to
    be able to reach a device that already has one."""
    versioned = client.get("/static/css/app.css?v=123").headers["Cache-Control"]
    assert "immutable" in versioned
    assert "max-age=31536000" in versioned

    plain = client.get("/static/icons/icon-32.png").headers["Cache-Control"]
    assert "immutable" not in plain

    # Photos sit behind the PIN, so they are not for a shared cache.
    assert "private" in client.get("/data/photos/nope.jpg?v=1").headers.get(
        "Cache-Control", "private"
    )


def test_a_missing_page_is_a_sentence_not_json(client, sample):
    """Staff used to get {"detail":"Not Found"} on a blank page, which reads as
    the app being broken rather than a link being stale. It also has to offer a
    way out, because the error page carries no nav bar."""
    response = client.get("/products/999999")
    assert response.status_code == 404
    assert "text/html" in response.headers["content-type"]
    assert "detail" not in response.text
    assert "Not found" in response.text
    assert 'href="/"' in response.text          # a way back


def test_the_error_page_needs_nothing_from_the_database(client, sample):
    """It does not extend base.html and asks for no settings row, so it still
    renders when the thing that is broken *is* the database."""
    raw = (ROOT / "app" / "templates" / "error.html").read_text(encoding="utf-8")
    # Comments first: the reasoning above names the very things being banned.
    template = re.sub(r"\{#.*?#\}", "", raw, flags=re.DOTALL)
    assert "extends" not in template
    assert "shop_name" not in template
    assert "{{ user" not in template


def test_a_redirect_raised_as_an_exception_is_still_a_redirect(client, sample, db, staff):
    """current_user bounces a retired account by raising a 303 with a Location
    header. The error-page handler must let that through — turning it into an
    HTML page saying "303" would quietly break the sign-in list."""
    db.execute("UPDATE users SET active = 0 WHERE id = ?", (staff["id"],))
    db.commit()

    response = client.post(
        f"/products/{sample['products']['monster']}/batches/"
        f"{sample['batches']['overdue']}",
        data={"status": "discounted"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_the_product_photo_opens_a_larger_view(client, sample, db):
    """Tapping the thumbnail opens the full-size photo. The thumbnail is marked
    for lightbox.js; a product with no photo has nothing to enlarge."""
    with_photo = sample["products"]["monster"]
    db.execute(
        "UPDATE products SET image_path = ? WHERE id = ?",
        ("data/photos/9300601234567.jpg", with_photo),
    )
    db.commit()

    body = client.get(f"/products/{with_photo}").text
    assert "data-lightbox" in body
    assert "lightbox.js" in body

    # The curly-apostrophe product has no photo — nothing to open.
    without = client.get(f"/products/{sample['products']['curly']}").text
    assert "data-lightbox" not in without


def test_product_list_filters_by_category(client, sample):
    body = client.get(f"/products?category={sample['cat_id']}").text
    assert "Monster Ultra Zero 500ml" in body
    assert "C/RIDGE WATER 1L" not in body


def test_product_list_sorts_by_soonest_expiry(client, sample):
    """Monster is 4 days overdue, the Tim Tams have a discounted date tomorrow,
    the water is due in 2.

    The Tim Tams moved ahead of the water on 13 Aug: their day-1 batch is
    `discounted` rather than `pulled`, so it counts as live and is now their
    soonest date. That is the status change working, not the sort breaking.
    """
    body = client.get("/products").text
    assert body.index("Monster Ultra Zero") < body.index("Tim Tam") < body.index("C/RIDGE WATER")


def test_product_detail_shows_every_batch_and_who_added_them(client, sample):
    body = client.get(f"/products/{sample['products']['curly']}").text
    assert _au(days(7)) in body        # active
    assert _au(days(1)) in body        # discounted — still on the shelf
    assert "Mohit" in body             # the audit trail, not the signed-in user


def test_product_detail_404s_for_an_unknown_id(client, sample):
    assert client.get("/products/999999").status_code == 404


def test_resolving_a_batch_records_who_and_when(client, sample, db, staff):
    batch = sample["batches"]["due_soon"]
    response = client.post(
        f"/products/{sample['products']['shouty']}/batches/{batch}",
        data={"status": "discounted"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    row = db.execute(
        "SELECT status, resolved_by, resolved_at FROM batches WHERE id = ?", (batch,)
    ).fetchone()
    assert row["status"] == "discounted"
    assert row["resolved_by"] == staff["id"]
    assert row["resolved_at"] is not None


def test_only_deleting_takes_a_batch_off_the_due_list(client, sample):
    """Rewritten 13 Aug. This used to mark the batch 'sold'.

    No status removes anything from the worklist any more — 'discounted' stays
    on it, because the item is still on the shelf. Deletion is the only way off,
    and it is a real delete.
    """
    batch = sample["batches"]["due_soon"]
    client.post(f"/products/{sample['products']['shouty']}/batches/{batch}/delete")
    assert "C/RIDGE WATER 1L" not in client.get("/").text


def test_a_discounted_batch_is_still_on_the_shelf(client, sample, db):
    """Discounted is a resolution but not a removal — it still shows as due."""
    batch = sample["batches"]["due_soon"]
    client.post(f"/products/{sample['products']['shouty']}/batches/{batch}",
                data={"status": "discounted"})
    assert "C/RIDGE WATER 1L" in client.get("/").text


def test_deleting_a_batch_really_deletes_it(client, sample, db):
    """The reverse of what this file asserted until 13 Aug.

    `test_nothing_is_hard_deleted` used to live here and guarded the opposite
    rule. The shop decided against keeping a record of everything ever removed:
    it only grows, on a laptop nobody prunes, and neither status that did it was
    ever used once. The Excel export is where a snapshot gets kept.
    """
    batch = sample["batches"]["due_soon"]
    before = db.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
    client.post(f"/products/{sample['products']['shouty']}/batches/{batch}/delete")
    assert db.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == before - 1
    assert db.execute("SELECT COUNT(*) FROM batches WHERE id = ?", (batch,)).fetchone()[0] == 0


def test_the_product_survives_deleting_its_last_batch(client, sample, db):
    """A product is never deleted. Not by staff, not at zero batches, not ever."""
    product = sample["products"]["shouty"]
    for batch in db.execute(
        "SELECT id FROM batches WHERE product_id = ?", (product,)
    ).fetchall():
        client.post(f"/products/{product}/batches/{batch['id']}/delete")

    assert db.execute("SELECT COUNT(*) FROM batches WHERE product_id = ?",
                      (product,)).fetchone()[0] == 0
    assert db.execute("SELECT COUNT(*) FROM products WHERE id = ?",
                      (product,)).fetchone()[0] == 1
    assert client.get(f"/products/{product}").status_code == 200


@pytest.mark.parametrize("status", ["binned", "pulled", "sold"])
def test_an_unknown_resolution_is_refused(client, sample, status):
    """'pulled' and 'sold' are unknown statuses now, not merely unused ones."""
    response = client.post(
        f"/products/{sample['products']['shouty']}/batches/{sample['batches']['due_soon']}",
        data={"status": status},
    )
    assert response.status_code == 400


def test_a_discounted_batch_can_go_back_to_full_price(client, sample, db):
    """Added 13 Aug: discounting is one tap, so people will do it to the wrong row.

    Not a deletion — the item is still on the shelf, just not marked down.
    """
    batch = sample["batches"]["discounted"]
    response = client.post(
        f"/products/{sample['products']['curly']}/batches/{batch}/full-price",
        follow_redirects=False,
    )
    assert response.status_code == 303

    row = db.execute(
        "SELECT status, resolved_by, resolved_at FROM batches WHERE id = ?", (batch,)
    ).fetchone()
    assert row["status"] == "active"
    # resolved_by means "who discounted this", so a batch nobody discounted has
    # nobody against it.
    assert row["resolved_by"] is None
    assert row["resolved_at"] is None


def test_going_back_to_full_price_keeps_the_batch(client, sample, db):
    """It is an undo of the discount, not of the batch."""
    batch = sample["batches"]["discounted"]
    before = db.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
    client.post(f"/products/{sample['products']['curly']}/batches/{batch}/full-price")
    assert db.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == before
    assert db.execute("SELECT COUNT(*) FROM batches WHERE id = ?",
                      (batch,)).fetchone()[0] == 1


def test_the_full_price_button_only_shows_on_a_discounted_row(client, sample):
    """An active batch has nothing to undo."""
    body = client.get("/").text
    assert body.count("/full-price") == 1        # only the discounted one


def test_only_a_discounted_batch_can_go_back_to_full_price(client, sample, db):
    """Guarded in the SQL, so this cannot quietly un-resolve something else."""
    batch = sample["batches"]["due_soon"]        # active already
    client.post(f"/products/{sample['products']['shouty']}/batches/{batch}/full-price")
    row = db.execute("SELECT status FROM batches WHERE id = ?", (batch,)).fetchone()
    assert row["status"] == "active"             # unchanged, and no error


def test_the_camera_does_not_save_by_itself(client):
    """From the iPad test: Take photo submitted the form and left edit mode.

    Taking a photo is one step of filling the form in, not the act of submitting
    it — it used to commit a name somebody might still have been typing.
    """
    import re

    js = client.get("/static/js/photo.js").text
    # Quoted, because 'camera-take' is a substring of 'camera-taken' and an
    # unquoted split lands on the wrong occurrence.
    handler = re.search(
        r"getElementById\('camera-take'\)\.addEventListener\('click',(.*?)\n  \}\);",
        js, re.S,
    )
    assert handler, "the camera-take handler is not in the shape expected"
    assert "upload(" not in handler.group(1), "taking a photo uploads immediately again"
    assert "captured = blob" in handler.group(1)
    assert "camera-taken" in js, "nothing tells the person the photo is held"


def test_the_product_screen_says_the_photo_is_waiting(client, sample):
    """Otherwise a taken photo looks like nothing happened."""
    body = client.get(f"/products/{sample['products']['monster']}").text
    assert 'id="camera-taken"' in body
    assert "press <strong>Save</strong>" in body


def test_a_plus_button_reaches_the_scan_page_from_everywhere(client, sample):
    """Adding a date happens hundreds of times a week — always under a thumb."""
    for path in ("/", "/products", "/sheet", "/settings",
                 f"/products/{sample['products']['monster']}"):
        body = client.get(path).text
        assert 'class="fab no-print"' in body, path
        assert 'href="/scan"' in body, path


def test_the_plus_button_is_not_on_the_printed_sheet(client, sample):
    """It is no-print, and the print CSS has to actually hide that class."""
    css = client.get("/static/css/app.css").text
    print_block = css.split("@page")[1]
    assert ".no-print" in print_block


def test_the_plus_button_is_hidden_from_anyone_not_signed_in(anon_client, staff):
    assert 'class="fab' not in anon_client.get("/login").text


def test_a_batch_date_can_be_corrected(client, sample, db, staff):
    """Punch list item 5. A date saved wrong used to live forever."""
    batch = sample["batches"]["due_soon"]
    response = client.post(
        f"/products/{sample['products']['shouty']}/batches/{batch}/date",
        data={"expiry_date": days(9)},
        follow_redirects=False,
    )
    assert response.status_code == 303

    row = db.execute(
        "SELECT expiry_date, edited_by, edited_at, added_by FROM batches WHERE id = ?",
        (batch,),
    ).fetchone()
    assert row["expiry_date"] == days(9)
    assert row["edited_by"] == staff["id"]
    assert row["edited_at"] is not None
    # Who first recorded it survives — that is what editing buys over a delete.
    assert row["added_by"] == sample["user_id"]


def test_correcting_a_date_onto_a_live_one_is_refused(client, sample, db):
    """The same duplicate check as an add, from app/catalogue.py.

    The Tim Tams have a live batch at +1 and another at +7. Moving the +7 onto
    +1 is the duplicate the app exists to prevent, and it must say so rather
    than surface an IntegrityError.
    """
    batch = sample["batches"]["due_edge"]          # curly, +7
    response = client.post(
        f"/products/{sample['products']['curly']}/batches/{batch}/date",
        data={"expiry_date": days(1)},             # curly already has this, discounted
    )
    assert response.status_code == 200
    assert "Already tracked" in response.text
    assert db.execute("SELECT expiry_date FROM batches WHERE id = ?",
                      (batch,)).fetchone()[0] == days(7), "the date moved anyway"


def test_saving_a_date_unchanged_is_not_a_duplicate_of_itself(client, sample, db):
    """Without excluding the row being edited, this finds itself and refuses."""
    batch = sample["batches"]["due_soon"]
    response = client.post(
        f"/products/{sample['products']['shouty']}/batches/{batch}/date",
        data={"expiry_date": days(2)},             # the date it already has
        follow_redirects=False,
    )
    assert response.status_code == 303, "a batch was refused as its own duplicate"
    assert db.execute("SELECT expiry_date FROM batches WHERE id = ?",
                      (batch,)).fetchone()[0] == days(2)


def test_a_corrected_date_still_refuses_an_implausible_year(client, sample, db):
    """Same parse_expiry as the add path — a mistyped year poisons the due list."""
    batch = sample["batches"]["due_soon"]
    response = client.post(
        f"/products/{sample['products']['shouty']}/batches/{batch}/date",
        data={"expiry_date": "2126-08-14"},
    )
    assert response.status_code == 200
    assert "Check the year" in response.text
    assert db.execute("SELECT expiry_date FROM batches WHERE id = ?",
                      (batch,)).fetchone()[0] == days(2)


def test_the_product_screen_shows_who_changed_a_date(client, sample, db):
    """Attribution is the point — it sits alongside who added it, not instead."""
    batch = sample["batches"]["due_soon"]
    client.post(f"/products/{sample['products']['shouty']}/batches/{batch}/date",
                data={"expiry_date": days(9)})
    body = client.get(f"/products/{sample['products']['shouty']}").text
    assert "date changed" in body
    assert STAFF_NAME in body            # who corrected it
    assert "Mohit" in body               # and who first recorded it


def test_the_product_screen_hides_its_editing_controls(client, sample):
    """Punch list item 4. The pickers used to sit on screen permanently, so a
    saved category read like unfinished work rather than a saved fact."""
    body = client.get(f"/products/{sample['products']['monster']}").text
    assert '<details class="edit"' in body
    assert "<summary" in body
    # Closed by default: <details> without `open` is collapsed.
    assert '<details class="edit" >' in body or '<details class="edit">' in body
    assert 'class="edit" open' not in body


def test_one_form_covers_name_category_and_photo(client, sample):
    """One toggle, one form, one Save — not three pickers with three Saves."""
    body = client.get(f"/products/{sample['products']['monster']}").text
    form = body.split('class="edit-form"')[1].split("</form>")[0]
    for field in ('id="name"', 'id="category"', 'id="photo"'):
        assert field in form, field
    assert form.count('type="submit"') == 1
    # The barcode is identity, not an editable field.
    assert 'name="barcode"' not in form


def test_a_product_can_be_renamed(client, sample, db):
    """Settles the question iteration 1 left open: staff can rename a product.

    A name typed wrong at 7am on a Saturday should be correctable. Nothing
    renames automatically and there is no bulk tidy-up.
    """
    product = sample["products"]["monster"]
    response = client.post(f"/products/{product}/edit",
                           data={"name": "Monster Ultra Zero 500mL", "category": ""},
                           follow_redirects=False)
    assert response.status_code == 303
    assert db.execute("SELECT name FROM products WHERE id = ?",
                      (product,)).fetchone()[0] == "Monster Ultra Zero 500mL"


def test_renaming_keeps_every_batch(client, sample, db):
    """The product is one row, so a rename carries its whole history with it."""
    product = sample["products"]["monster"]
    before = db.execute("SELECT COUNT(*) FROM batches WHERE product_id = ?",
                        (product,)).fetchone()[0]
    client.post(f"/products/{product}/edit", data={"name": "Renamed Thing"})
    assert db.execute("SELECT COUNT(*) FROM batches WHERE product_id = ?",
                      (product,)).fetchone()[0] == before


def test_a_blank_name_is_refused_and_reopens_the_panel(client, sample, db):
    """Never silently keep the old name — and never close on what was typed."""
    product = sample["products"]["monster"]
    response = client.post(f"/products/{product}/edit", data={"name": "   "})
    assert response.status_code == 200
    assert "Give the product a name" in response.text
    assert 'class="edit" open' in response.text
    assert db.execute("SELECT name FROM products WHERE id = ?",
                      (product,)).fetchone()[0] == MONSTER_NAME


def test_a_messy_imported_name_is_left_alone_by_default(client, sample, db):
    """Trailing whitespace and curly apostrophes survive an unrelated edit.

    Staff recognise these names as they are. Editing the category must not
    quietly tidy the name — see docs/DATA-NOTES.md.
    """
    product = sample["products"]["shouty"]
    original = db.execute("SELECT name FROM products WHERE id = ?",
                          (product,)).fetchone()[0]
    assert original == "C/RIDGE WATER 1L  "        # the real export has this

    body = client.get(f"/products/{product}").text
    assert 'value="C/RIDGE WATER 1L  "' in body, "the field must offer it verbatim"


def test_the_edit_panel_reopens_on_a_bad_photo(client, sample):
    """A rejected save must not close the panel and swallow the typing."""
    response = _edit(
        client, sample["products"]["monster"],
        files={"photo": ("notes.txt", b"not a photograph", "text/plain")},
    )
    assert "not an image" in response.text
    assert 'class="edit" open' in response.text


def test_a_bad_photo_does_not_save_the_name_either(client, sample, db):
    """One form, one save: if it is refused, none of it lands."""
    product = sample["products"]["monster"]
    _edit(client, product, name="Should Not Persist",
          files={"photo": ("notes.txt", b"not a photograph", "text/plain")})
    assert db.execute("SELECT name FROM products WHERE id = ?",
                      (product,)).fetchone()[0] == MONSTER_NAME


def test_the_photo_script_sends_the_whole_form(client):
    """It used to post a FormData holding only the photo.

    That was right when the photo had a route of its own. Now name, category and
    photo save together, so sending just the image would discard a rename typed
    in the same panel — and the shrink path is the one an iPad always takes.

    Guaranteed structurally since 16 Aug rather than by assembling a FormData
    carefully: the script puts the shrunk image back into the file input and
    lets the browser submit the form, so every field on it goes by definition
    and there is no second copy of the form to keep in step.
    """
    js = client.get("/static/js/photo.js").text
    assert "input.files = carrier.files" in js, "the shrunk image is not put back"
    assert "form.submit()" in js, "the form is not submitted by the browser"
    assert 'button[type="submit"]' in js, "a bare button selector grabs Camera"


def test_category_can_be_set_from_the_product_screen(client, sample, db):
    client.post(f"/products/{sample['products']['shouty']}/edit",
                data={"name": "C/RIDGE WATER 1L", "category": "Water"})
    assert db.execute(
        """SELECT c.name FROM products p JOIN categories c ON c.id = p.category_id
            WHERE p.id = ?""",
        (sample["products"]["shouty"],),
    ).fetchone()[0] == "Water"


def test_clearing_the_category_is_allowed(client, sample, db):
    client.post(f"/products/{sample['products']['monster']}/edit",
                data={"name": MONSTER_NAME, "category": ""})
    assert db.execute(
        "SELECT category_id FROM products WHERE id = ?", (sample["products"]["monster"],)
    ).fetchone()[0] is None


# ----------------------------------------------------------------- photos

def _photo_bytes(width=2000, height=1500, orientation=None) -> bytes:
    """A JPEG the size an iPad actually produces, optionally rotated by EXIF."""
    import io

    from PIL import Image

    image = Image.new("RGB", (width, height), (180, 40, 40))
    for x in range(0, width, 40):           # some detail, so it isn't a flat block
        for y in range(0, height, 40):
            image.putpixel((x, y), (20, 90, 200))
    buffer = io.BytesIO()
    if orientation:
        exif = image.getexif()
        exif[274] = orientation
        image.save(buffer, "JPEG", exif=exif, quality=95)
    else:
        image.save(buffer, "JPEG", quality=95)
    return buffer.getvalue()


# Every product edit goes through one form now, and `name` is required — so the
# photo tests carry the existing name along rather than each asserting a rename.
MONSTER_NAME = "Monster Ultra Zero 500ml"


def _edit(client, product_id, name=MONSTER_NAME, **kwargs):
    """POST the combined edit form. kwargs go straight through (files=, data=)."""
    data = {"name": name}
    data.update(kwargs.pop("data", {}))
    return client.post(f"/products/{product_id}/edit", data=data, **kwargs)


def test_uploading_a_photo_attaches_it_to_the_product(client, sample, db, photo_dir):
    response = _edit(
        client, sample["products"]["monster"],
        files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")},
        follow_redirects=False,
        )
    assert response.status_code == 303

    stored = db.execute(
        "SELECT image_path FROM products WHERE id = ?", (sample["products"]["monster"],)
    ).fetchone()[0]
    assert stored == "data/photos/9300601234567.jpg"     # keyed to the barcode
    assert (photo_dir / "9300601234567.jpg").exists()


def test_a_photo_is_resized_and_stripped(client, sample, photo_dir):
    from PIL import Image

    _edit(
        client, sample["products"]["monster"],
        files={"photo": ("counter.jpg", _photo_bytes(orientation=6), "image/jpeg")},
    )
    saved = Image.open(photo_dir / "9300601234567.jpg")

    assert saved.format == "JPEG"
    assert max(saved.size) == 800                        # long edge, SPEC §5
    assert saved.size == (600, 800), "EXIF orientation 6 means this is portrait"
    assert not dict(saved.getexif()), "EXIF is stripped — an iPad photo carries GPS"
    assert (photo_dir / "9300601234567.jpg").stat().st_size < 80 * 1024


def test_an_uploaded_photo_actually_serves(client, sample, db, photo_dir):
    """Punch list item 8, and the test that could not exist before it.

    The photo folder was served by a StaticFiles mount over a hardcoded path
    while uploads honoured TTD_PHOTO_DIR, so with the variable set the two
    disagreed and every image rendered broken — which is what the iPad session
    saw. A mount binds its directory once at import, and the suite points each
    test at a fresh temp folder, so no test could ever have caught it. Served by
    a route resolved per request, this fails when they disagree.
    """
    _edit(client, sample["products"]["monster"],
          files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")})

    stored = db.execute(
        "SELECT image_path FROM products WHERE id = ?", (sample["products"]["monster"],)
    ).fetchone()[0]
    assert stored == "data/photos/9300601234567.jpg"

    response = client.get(f"/{stored}")
    assert response.status_code == 200, "the photo is stored but does not serve"
    assert response.content[:2] == b"\xff\xd8", "that is not a JPEG"
    assert len(response.content) == (photo_dir / "9300601234567.jpg").stat().st_size


def test_a_photo_url_on_the_page_can_be_fetched(client, sample, db):
    """The <img> src the Due screen renders has to be a URL that works."""
    import re

    _edit(client, sample["products"]["monster"],
          files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")})

    src = re.search(r'<img class="thumb" src="([^"]+)"', client.get("/").text)
    assert src, "no photo rendered on the Due screen"
    assert client.get(src.group(1)).status_code == 200


def test_a_photo_request_cannot_climb_out_of_the_folder(client):
    """The one place a request names a file on disk."""
    for attempt in ("../../app/schema.sql", "..%2f..%2fapp%2fschema.sql", "../tecoma.db"):
        assert client.get(f"/data/photos/{attempt}").status_code in (404, 400), attempt


def test_a_photo_appears_on_batches_recorded_months_ago(client, sample):
    """It hangs off the barcode, so it backfills everything already recorded."""
    assert "thumb-empty" in client.get("/").text
    _edit(
        client, sample["products"]["monster"],
        files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")},
    )
    body = client.get("/").text
    assert "/data/photos/9300601234567.jpg" in body


def test_replacing_a_photo_leaves_no_orphan(client, sample, photo_dir):
    for _ in range(3):
        _edit(
            client, sample["products"]["monster"],
            files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")},
        )
    files = [p for p in photo_dir.iterdir() if p.is_file()]
    assert len(files) == 1, files


def test_the_photo_url_changes_when_the_photo_does(client, sample):
    """Same filename every time, so without a stamp an iPad shows a stale one."""
    _edit(
        client, sample["products"]["monster"],
        files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")},
    )
    body = client.get(f"/products/{sample['products']['monster']}").text
    assert "9300601234567.jpg?v=" in body
    assert "?v=0" not in body


def test_a_file_that_is_not_an_image_is_refused_politely(client, sample, db):
    response = _edit(
        client, sample["products"]["monster"],
        files={"photo": ("notes.txt", b"this is not a photograph", "text/plain")},
    )
    assert response.status_code == 200
    assert "not an image" in response.text
    assert db.execute(
        "SELECT image_path FROM products WHERE id = ?", (sample["products"]["monster"],)
    ).fetchone()[0] is None


def test_the_photo_form_offers_a_file_and_a_camera(client, sample):
    """The file input covers uploads and Take Photo on iOS; the camera button
    is unhidden by script only where getUserMedia can actually work."""
    body = client.get(f"/products/{sample['products']['monster']}").text
    assert 'type="file"' in body and 'accept="image/*"' in body
    assert 'id="camera"' in body
    js = client.get("/static/js/photo.js").text
    assert "getUserMedia" in js
    assert "canvas" in js               # shrunk in the browser before upload


def test_a_product_with_no_photo_is_a_normal_state(client, sample):
    """Nothing in the add path waits for a camera."""
    for path in ("/", "/products", f"/products/{sample['products']['curly']}"):
        assert "thumb-empty" in client.get(path).text


# --------------------------------------------------- weekly discount sheet

def test_sheet_lists_what_is_due_in_the_window(client, sample, db):
    body = client.get("/sheet").text
    assert "C/RIDGE WATER 1L" in body                  # due in 2 days
    far_off = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["outside"],)
    ).fetchone()[0]
    assert _au(far_off) not in body                    # 30 days out, not this week


def test_sheet_range_is_adjustable(client, sample, db):
    far_off = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["outside"],)
    ).fetchone()[0]
    assert _au(far_off) in client.get("/sheet?days=45").text


def test_sheet_groups_by_category_with_uncategorised_last(client, sample, db):
    """The categorised item needs an in-window date now the range is bounded."""
    db.execute(
        "INSERT INTO batches (product_id, expiry_date, status, added_by) "
        "VALUES (?, ?, 'active', ?)",
        (sample["products"]["monster"], days(3), sample["user_id"]),
    )
    db.commit()
    body = client.get("/sheet").text
    assert body.index("Energy Drinks") < body.index("No category")
    assert "Uncategorised" not in body


def test_sheet_sorts_by_date_within_a_group(client, sample, db):
    """The uncategorised group, in date order.

    Asserted on the rendered dates rather than on two product names: the Tim
    Tams now hold both the soonest date in the group (day 1, discounted) and the
    latest (day 7), so naming products pins the test to which batch happens to
    come first rather than to the ordering itself.
    """
    import re

    body = client.get("/sheet").text
    group = body.split("No category")[1]
    dates = re.findall(r'<td class="col-date">([^<]+)</td>', group)
    assert len(dates) >= 2, "expected several rows in the uncategorised group"
    parsed = [_from_au(d) for d in dates]
    assert parsed == sorted(parsed), f"out of order: {dates}"


def test_sheet_has_a_tick_box_and_a_blank_price_column(client, sample):
    body = client.get("/sheet").text
    assert 'class="tick"' in body
    assert "Price" in body
    assert '<td class="col-price"></td>' in body       # blank, for a pen


def test_sheet_shows_the_barcode_on_every_line(client, sample):
    body = client.get("/sheet").text
    assert "9300609876543" in body


def test_sheet_shows_a_discounted_batch(client, sample, db):
    """Inverted 13 Aug with the status change — see test_home_shows_a_discounted_batch.

    A discounted item is still sellable and still wants a sticker checked, so it
    belongs on the sheet. What the sheet leaves out is past dates, which is a
    question about the date rather than the status.
    """
    discounted = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["discounted"],)
    ).fetchone()[0]
    assert _au(discounted) in client.get("/sheet").text


def test_sheet_leaves_out_past_date_items(client, sample, db):
    """Amended 13 Aug: the sheet is a pricing list, not the worklist.

    A past-date item is not discounted, it is pulled off the shelf. Before this
    the shop's real sheet opened with 27 of them ahead of the week's work.
    """
    overdue = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["overdue"],)
    ).fetchone()[0]
    body = client.get("/sheet").text
    assert _au(overdue) not in body
    assert "Monster Ultra Zero 500ml" not in body   # its only other date is 30 days out
    assert "(past)" not in body                     # the marker has nothing left to mark


def test_the_due_screen_still_shows_the_past_date_backlog(client, sample, db):
    """The two screens diverge deliberately — this is the half that must not.

    One definition of "due" survives: the Due screen is the worklist, and a
    past-date item is the most urgent thing on it. If bounding the sheet ever
    gets copied into home.py, this goes red.
    """
    overdue = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["overdue"],)
    ).fetchone()[0]
    body = client.get("/").text
    assert _au(overdue) in body
    assert "Monster Ultra Zero 500ml" in body


def test_sheet_includes_an_item_expiring_today(client, sample, db):
    """The lower bound is inclusive: today's stock still wants a sticker.

    Asserted on a product name rather than on today's date, because the sheet
    title prints today either way — checking the date alone passes against a
    sheet bounded with `>` and catches nothing.
    """
    product_id = db.execute(
        "INSERT INTO products (barcode, name) VALUES ('9300602222222', ?)",
        ("Dies Today Yoghurt 170g",),
    ).lastrowid
    db.execute(
        "INSERT INTO batches (product_id, expiry_date, status, added_by) "
        "VALUES (?, ?, 'active', ?)",
        (product_id, days(0), sample["user_id"]),
    )
    db.commit()
    assert "Dies Today Yoghurt 170g" in client.get("/sheet").text


def test_sheet_survives_an_absurd_range(client, sample):
    assert client.get("/sheet?days=99999").status_code == 200
    assert client.get("/sheet?days=0").status_code == 200


def test_sheet_prints_without_the_navigation(client, sample):
    """A4 print CSS: no nav, no chrome, no toner-eating backgrounds."""
    css = client.get("/static/css/app.css").text
    assert "@page" in css and "size: A4" in css
    assert ".bar, .no-print" in css


# --------------------------------------------------------------- settings

def test_settings_opens_for_anyone_signed_in(client):
    """No admin tier, no manager PIN — everyone gets the same app."""
    response = client.get("/settings")
    assert response.status_code == 200
    for word in ("Expiry window", "Categories", "Staff", "Backup"):
        assert word in response.text


def test_the_expiry_window_can_be_changed_and_the_screens_follow(client, sample, db):
    """It is a setting precisely so this does not need a code change."""
    edge = db.execute(
        "SELECT expiry_date FROM batches WHERE id = ?", (sample["batches"]["due_edge"],)
    ).fetchone()[0]
    assert _au(edge) in client.get("/").text           # 7 days out, inside the window

    client.post("/settings/window", data={"days": 3})
    assert db.execute(
        "SELECT value FROM settings WHERE key = 'expiry_window_days'"
    ).fetchone()[0] == "3"

    # Changed 16 Aug with the day bands. The window no longer decides what is on
    # this screen — it decides where the per-day bands stop. The 7-day item is
    # still on the page with a 3-day window; it has moved out of the bands and
    # under "More than 3 days".
    body = client.get("/").text
    assert _au(edge) in body
    assert "More than 3 days" in body
    assert "More than a week" not in body       # the wording follows the window


def test_an_absurd_window_is_clamped_not_crashed(client, db):
    client.post("/settings/window", data={"days": 100000})
    assert client.get("/").status_code == 200


def test_a_category_can_be_deleted(client, sample, db):
    """Punch list item 6. A typo in the filter list was permanent before this."""
    response = client.post(f"/settings/categories/{sample['cat_id']}/delete",
                           follow_redirects=False)
    assert response.status_code == 303
    assert db.execute("SELECT COUNT(*) FROM categories WHERE id = ?",
                      (sample["cat_id"],)).fetchone()[0] == 0


def test_deleting_a_category_keeps_its_products(client, sample, db):
    """ON DELETE SET NULL. Losing a category must never lose the catalogue."""
    product = sample["products"]["monster"]
    client.post(f"/settings/categories/{sample['cat_id']}/delete")

    row = db.execute("SELECT category_id FROM products WHERE id = ?", (product,)).fetchone()
    assert row is not None, "the product went with the category"
    assert row["category_id"] is None       # uncategorised, which is normal
    assert client.get(f"/products/{product}").status_code == 200
    assert client.get("/").status_code == 200


def test_deleting_a_category_names_it_and_says_what_uses_it(client, sample, db):
    """Reworded 13 Aug after the iPad test: the question names the category.

    "Delete Confectionary" with a separate "Nothing is using it" read as two
    unrelated bits of text. It is one sentence now, and the number is the whole
    point of it — one product is nothing, ninety is worth a second look.
    """
    # Whitespace-collapsed: the sentence spans several lines in the template, and
    # asserting on the raw HTML would make it a test of the indentation.
    body = " ".join(client.get("/settings").text.split())
    assert "Are you sure you want to delete Energy Drinks?" in body
    assert "1 product uses it, and it becomes uncategorised" in body

    db.execute("INSERT INTO products (barcode, name, category_id) VALUES (?, ?, ?)",
               ("9300603333333", "Another Energy Drink", sample["cat_id"]))
    db.commit()
    body = " ".join(client.get("/settings").text.split())
    assert "2 products use it, and they become uncategorised" in body


def test_the_category_delete_sits_beside_rename(client, sample):
    """It used to be on its own line below, reading like a separate section."""
    body = client.get("/settings").text
    row = body.split('class="cat-row"')[1].split("</details>")[0]
    assert ">Rename<" in row
    assert ">Delete<" in row


def test_an_unused_category_says_nothing_is_using_it(client, db, staff):
    """A fresh category has no products, and the sentence has to still read."""
    client.post("/settings/categories", data={"name": "Confectionary"})
    body = " ".join(client.get("/settings").text.split())
    assert "Are you sure you want to delete Confectionary? Nothing is using it." in body


def test_a_category_delete_is_not_gated_behind_a_role(client, sample, db):
    """No roles, by decision. Anyone signed in can do this."""
    assert "role" not in [c[1] for c in db.execute("PRAGMA table_info(users)")]
    assert client.post(f"/settings/categories/{sample['cat_id']}/delete",
                       follow_redirects=False).status_code == 303


def test_deleting_a_category_removes_it_from_every_filter(client, sample):
    """It is the filter list this exists to clean up."""
    assert "Energy Drinks" in client.get("/").text
    client.post(f"/settings/categories/{sample['cat_id']}/delete")
    for path in ("/", "/products", "/scan", "/settings"):
        assert "Energy Drinks" not in client.get(path).text, path


def test_a_category_can_be_added_here(client, db, staff):
    client.post("/settings/categories", data={"name": "Chocolate"})
    row = db.execute("SELECT created_by FROM categories WHERE name = 'Chocolate'").fetchone()
    assert row is not None and row["created_by"] == staff["id"]


def test_adding_a_case_variant_category_is_refused_politely(client, sample):
    body = client.post(
        "/settings/categories", data={"name": "ENERGY DRINKS"}, follow_redirects=True
    ).text
    assert "already a category with that name" in body
    assert "IntegrityError" not in body


def test_renaming_a_category_reaches_every_product_in_it(client, sample, db):
    client.post(f"/settings/categories/{sample['cat_id']}", data={"name": "Energy"})
    assert db.execute(
        """SELECT c.name FROM products p JOIN categories c ON c.id = p.category_id
            WHERE p.id = ?""",
        (sample["products"]["monster"],),
    ).fetchone()[0] == "Energy"


def test_staff_can_be_added_and_can_then_sign_in(client, anon_client, db):
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})
    new_id = db.execute("SELECT id FROM users WHERE name = 'Sarah'").fetchone()[0]

    client.get("/logout")
    response = anon_client.post(
        "/login", data={"user_id": new_id, "pin": "4821"}, follow_redirects=False
    )
    assert response.status_code == 303


def test_a_pin_that_is_not_four_digits_is_refused(client, db):
    body = client.post(
        "/settings/staff", data={"name": "Bad PIN", "pin": "12"}, follow_redirects=True
    ).text
    assert "exactly 4 digits" in body
    assert db.execute("SELECT COUNT(*) FROM users WHERE name = 'Bad PIN'").fetchone()[0] == 0


def test_resetting_a_pin_changes_which_one_works(client, anon_client, staff):
    client.post(f"/settings/staff/{staff['id']}/pin", data={"pin": "5150"})
    client.get("/logout")

    assert anon_client.post(
        "/login", data={"user_id": staff["id"], "pin": STAFF_PIN}, follow_redirects=False
    ).status_code == 200                                  # the old one no longer works
    assert anon_client.post(
        "/login", data={"user_id": staff["id"], "pin": "5150"}, follow_redirects=False
    ).status_code == 303


# ------------------------------------------------------ camera scanning
# Iteration 2 item 3. The aisle half of scanning. The counter gun is the path
# used hundreds of times a week and must come out of this untouched.

def test_the_camera_button_ships_hidden(client):
    """It is revealed by JS only where getUserMedia exists.

    Anywhere a camera could not be opened — the shop's http address before the
    certificates go on — it stays hidden rather than appearing and failing.
    """
    body = client.get("/scan").text
    tag = body.split('id="scan-camera"')[1].split(">")[0]
    assert "hidden" in tag


def test_the_camera_fills_the_same_field_and_submits_the_same_form(client):
    """One route, one lookup, one duplicate check — however the barcode arrived."""
    body = client.get("/scan").text
    assert 'id="barcode-form"' in body
    assert 'action="/scan"' in body
    assert 'method="get"' in body
    # The button is inside that form and is type=button, so it cannot itself
    # submit a blank lookup before anything has been decoded.
    tag = body.split('id="scan-camera"')[1].split(">")[0]
    assert 'type="button"' in tag


def test_a_decoded_barcode_takes_the_ordinary_path(client, sample, db, staff):
    """Whatever the camera fills in is just a barcode in the query string."""
    body = client.get("/scan?barcode=9300601234567").text
    assert "Monster Ultra Zero 500ml" in body

    # ...and the duplicate rule still applies to it, exactly as when typed.
    client.post("/scan/add", data={"barcode": "9300601234567", "expiry_date": days(50)})
    body = client.post(
        "/scan/add", data={"barcode": "9300601234567", "expiry_date": days(50)},
        follow_redirects=True,
    ).text
    assert "Already tracked" in body
    assert db.execute(
        "SELECT COUNT(*) FROM batches WHERE product_id = ? AND expiry_date = ?",
        (sample["products"]["monster"], days(50)),
    ).fetchone()[0] == 1


def test_the_scanner_library_is_served_from_static_vendor(client):
    response = client.get("/static/vendor/zxing-0.21.3.min.js")
    assert response.status_code == 200
    assert "javascript" in response.headers["content-type"]
    assert "ZXing" in response.text[:400]


def test_the_scanner_script_is_absent_from_the_date_step(client, sample):
    """The counter path is scan, type date, Enter. It carries no scanner code."""
    body = client.get("/scan?barcode=9300601234567").text
    assert "scanner.js" not in body
    assert 'id="scan-camera"' not in body


def test_the_heavy_library_is_not_loaded_up_front(client):
    """336 KB must not land on the gun path. scanner.js injects it on demand."""
    body = client.get("/scan").text
    assert "scanner.js" in body
    assert "zxing" not in body.lower()


def test_the_counter_flow_is_unchanged(client, sample, db, staff):
    """Scan, type date, Enter — no mouse, and nothing new in the way."""
    body = client.get("/scan").text
    assert "autofocus" in body.split('id="barcode"')[1].split(">")[0]

    body = client.get("/scan?barcode=9300601234567").text
    assert "autofocus" in body.split('id="expiry"')[1].split(">")[0]

    response = client.post(
        "/scan/add", data={"barcode": "9300601234567", "expiry_date": days(55)},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert db.execute(
        "SELECT added_by FROM batches WHERE expiry_date = ?", (days(55),)
    ).fetchone()[0] == staff["id"]


# ------------------------------------------------- renaming and retiring staff
# Iteration 2 item 1. The two imported accounts are `BP TECOMA` and `sar ob`,
# both still on the placeholder PIN, and until they are dealt with the audit
# trail says BP TECOMA for everybody.

def test_somebody_can_be_renamed(client, anon_client, db, staff):
    """`sar ob` becoming a real name is the whole point of this."""
    client.post(f"/settings/staff/{staff['id']}/name", data={"name": "Sarah O’Brien"})
    assert db.execute(
        "SELECT name FROM users WHERE id = ?", (staff["id"],)
    ).fetchone()[0] == "Sarah O’Brien"

    client.get("/logout")
    assert "Sarah O’Brien" in anon_client.get("/login").text
    assert STAFF_NAME not in anon_client.get("/login").text


def test_a_rename_carries_every_entry_that_person_ever_made(client, sample):
    """added_by points at the row, so the history follows the name.

    This is the behaviour that makes renaming the shared `BP TECOMA` login
    into a person wrong, and it is why the screen shows the entry count.
    """
    client.post(f"/settings/staff/{sample['user_id']}/name", data={"name": "Mohit Pandya"})
    body = client.get(f"/products/{sample['products']['monster']}").text
    assert "Added by Mohit Pandya" in body


def test_a_rename_does_not_move_any_history(client, sample, db):
    """Renaming must not touch the batches themselves — same rows, same ids."""
    before = db.execute(
        "SELECT COUNT(*) FROM batches WHERE added_by = ?", (sample["user_id"],)
    ).fetchone()[0]
    client.post(f"/settings/staff/{sample['user_id']}/name", data={"name": "Mohit Pandya"})
    assert db.execute(
        "SELECT COUNT(*) FROM batches WHERE added_by = ?", (sample["user_id"],)
    ).fetchone()[0] == before
    assert db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 2


def test_the_header_shows_the_new_name_without_signing_in_again(client, staff):
    """The cookie still holds the old name; the page must not."""
    client.post(f"/settings/staff/{staff['id']}/name", data={"name": "Sarah O’Brien"})
    body = client.get("/").text
    assert "Sarah O’Brien" in body
    assert STAFF_NAME not in body


def test_renaming_somebody_to_a_case_variant_of_somebody_else_is_refused(client, db, staff):
    """Two Sarahs on the keypad screen and nobody knows which one is theirs."""
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})
    body = client.post(
        f"/settings/staff/{staff['id']}/name", data={"name": "SARAH"}, follow_redirects=True
    ).text
    assert "already has that name" in body
    assert "IntegrityError" not in body
    assert db.execute(
        "SELECT name FROM users WHERE id = ?", (staff["id"],)
    ).fetchone()[0] == STAFF_NAME


def test_adding_a_case_variant_of_an_existing_person_is_refused(client, db, staff):
    body = client.post(
        "/settings/staff", data={"name": STAFF_NAME.lower(), "pin": "4821"},
        follow_redirects=True,
    ).text
    assert "already has that name" in body
    assert db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 1


def test_somebody_can_be_recased_without_colliding_with_themselves(client, db, staff):
    """`sar ob` -> `Sar Ob` is a rename, not a duplicate."""
    client.post(f"/settings/staff/{staff['id']}/name", data={"name": STAFF_NAME.upper()})
    assert db.execute(
        "SELECT name FROM users WHERE id = ?", (staff["id"],)
    ).fetchone()[0] == STAFF_NAME.upper()


def test_a_blank_staff_name_is_refused(client, db, staff):
    body = client.post(
        f"/settings/staff/{staff['id']}/name", data={"name": "   "}, follow_redirects=True
    ).text
    assert "Give the person a name" in body
    assert db.execute(
        "SELECT name FROM users WHERE id = ?", (staff["id"],)
    ).fetchone()[0] == STAFF_NAME


def test_somebody_taken_off_the_list_cannot_sign_in(client, anon_client, staff):
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})
    client.post(f"/settings/staff/{staff['id']}/active", data={"active": "0"})
    client.get("/logout")

    assert STAFF_NAME not in anon_client.get("/login").text
    assert anon_client.post(
        "/login", data={"user_id": staff["id"], "pin": STAFF_PIN}, follow_redirects=False
    ).status_code == 200                                  # re-rendered, not signed in


def test_taking_somebody_off_the_list_deletes_nothing(client, sample, db):
    """Their entries have to stay readable — that is the whole reason for it."""
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})
    client.post(f"/settings/staff/{sample['user_id']}/active", data={"active": "0"})

    assert db.execute(
        "SELECT COUNT(*) FROM batches WHERE added_by = ?", (sample["user_id"],)
    ).fetchone()[0] == 5
    assert "Added by Mohit" in client.get(
        f"/products/{sample['products']['monster']}"
    ).text


def test_somebody_taken_off_the_list_can_be_put_back(client, anon_client, db):
    """A misclick has to be undoable, so there is no confirm step on the way in."""
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})
    sarah = db.execute("SELECT id FROM users WHERE name = 'Sarah'").fetchone()[0]

    client.post(f"/settings/staff/{sarah}/active", data={"active": "0"})
    assert "Off the sign-in list" in client.get("/settings").text
    assert db.execute("SELECT active FROM users WHERE id = ?", (sarah,)).fetchone()[0] == 0

    client.post(f"/settings/staff/{sarah}/active", data={"active": "1"})
    assert "Off the sign-in list" not in client.get("/settings").text

    client.get("/logout")                      # the sign-in list needs a signed-out client
    assert "Sarah" in anon_client.get("/login").text


def test_the_staff_list_shows_how_many_entries_each_person_has(client, sample, staff):
    """So nobody renames a 2,290-entry shared login without seeing the 2,290."""
    body = client.get("/settings").text
    assert "5 entries" in body                            # sample's user
    assert "0 entries" in body                            # the signed-in one


def test_the_off_the_list_card_is_absent_when_everyone_is_on_it(client, staff):
    assert "Off the sign-in list" not in client.get("/settings").text


def test_backup_runs_on_demand_and_is_reported(client, db_path):
    body = client.get("/settings").text
    assert "No backup has been taken yet" in body

    client.post("/settings/backup")
    snapshots = list((db_path.parent / "backups").glob("tecoma-*.db"))
    assert len(snapshots) == 1
    assert "Last backup:" in client.get("/settings").text


def test_backup_snapshots_the_database_the_app_is_using(client, db_path, sample):
    """Never data/tecoma.db while a test is running — it follows the connection."""
    from scripts.init_db import DB_PATH, connect

    client.post("/settings/backup")
    snapshot = next((db_path.parent / "backups").glob("tecoma-*.db"))
    assert snapshot.parent != DB_PATH.parent

    copy = connect(snapshot)
    assert copy.execute("SELECT COUNT(*) FROM products").fetchone()[0] == 3
    copy.close()


# ----------------------------------------------------------- the Excel export

def _sheet(response):
    """The one worksheet in a downloaded export."""
    from io import BytesIO

    from openpyxl import load_workbook

    return load_workbook(BytesIO(response.content)).active


def test_settings_offers_the_export(client, sample):
    body = client.get("/settings").text
    assert "/settings/export.xlsx" in body
    assert "Export to Excel" in body
    assert "5 of them" in body                    # sample has five batches


def test_the_export_downloads_as_a_spreadsheet(client, sample):
    import datetime as dt

    response = client.get("/settings/export.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers["content-disposition"]
    assert "attachment" in disposition
    assert f"tecoma-{dt.date.today().isoformat()}.xlsx" in disposition


def test_the_export_has_one_row_per_batch_and_a_header(client, sample):
    sheet = _sheet(client.get("/settings/export.xlsx"))
    assert sheet.max_row == 6                      # header + five batches
    headings = [c.value for c in sheet[1]]
    assert headings[:6] == ["Product", "Barcode", "Category", "Expiry",
                            "Days left", "Status"]


def test_the_export_includes_discounted_batches(client, sample):
    """Every batch, and after 13 Aug that means every row that exists.

    A deleted batch cannot appear here, which is why the export has to be taken
    before a clear-out rather than after — it is the only record that keeps one.
    """
    sheet = _sheet(client.get("/settings/export.xlsx"))
    statuses = [row[5].value for row in sheet.iter_rows(min_row=2)]
    assert "discounted" in statuses
    assert statuses.count("active") == 4
    assert "pulled" not in statuses and "sold" not in statuses


def test_the_export_is_in_date_order(client, sample):
    """Soonest first, the same as every screen.

    The old sort lifted live rows above resolved ones, because 583 already-pulled
    rows were the oldest in the file and put the manager's cursor on the part that
    no longer mattered. Those rows are gone, so plain date order is the useful
    order again — and every remaining row is on the shelf.
    """
    dates = [row[3].value for row in _sheet(client.get("/settings/export.xlsx"))
             .iter_rows(min_row=2)]
    assert dates == sorted(dates), "the export is no longer in date order"


def test_the_export_shows_uncategorised_as_blank_not_a_word(client, sample):
    """There is no 'Uncategorised' category and the export must not invent one."""
    sheet = _sheet(client.get("/settings/export.xlsx"))
    categories = [row[2].value for row in sheet.iter_rows(min_row=2)]
    assert None in categories
    assert "Uncategorised" not in str(categories)


def test_the_export_counts_days_from_today(client, sample):
    sheet = _sheet(client.get("/settings/export.xlsx"))
    by_days = {row[4].value for row in sheet.iter_rows(min_row=2)}
    assert -4 in by_days                            # sample's overdue batch
    assert 7 in by_days                             # the one on the window edge


def test_the_export_names_who_added_each_row(client, sample):
    sheet = _sheet(client.get("/settings/export.xlsx"))
    assert {row[7].value for row in sheet.iter_rows(min_row=2)} == {"Mohit"}


def test_the_export_is_behind_the_pin(anon_client, sample):
    response = anon_client.get("/settings/export.xlsx", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"].startswith("/login")


def test_settings_has_no_role_or_admin_anything(client):
    body = client.get("/settings").text.lower()
    for word in ("admin", "manager", "role", "permission"):
        assert word not in body


def test_static_assets_are_served(client):
    response = client.get("/static/css/app.css")
    assert response.status_code == 200
    assert "text/css" in response.headers["content-type"]


# ------------------------------------------------- the home screen button
# The app is opened from an icon on an iPad's home screen, not by typing an
# address. Everything below is what makes that icon a button rather than a
# blurry screenshot of whichever page happened to be open when it was added.

def test_the_home_screen_icon_exists_and_is_linked(anon_client):
    """No apple-touch-icon and iOS uses a screenshot of the page as the button.

    Signed out on purpose: an iPad is added to the home screen before anyone
    signs in on it, so the icon has to be fetchable without a session.
    """
    body = anon_client.get("/login").text
    assert '<link rel="apple-touch-icon" href="/static/icons/icon-180.png">' in body

    response = anon_client.get("/static/icons/icon-180.png")
    assert response.status_code == 200
    assert response.headers["content-type"] == "image/png"


def test_the_icon_is_labelled_with_the_app_and_not_the_page(anon_client):
    """Without this iOS names the button after the page title, so the same
    app lands on ten iPads as Due, Scan, Sign in and Products."""
    assert '<meta name="apple-mobile-web-app-title" content="Track the Date">' in (
        anon_client.get("/login").text
    )


def test_the_android_manifest_is_reachable_and_is_json(anon_client):
    """The scanner phone's half of the same thing. Under /static because a
    manifest is fetched without the session cookie — behind the middleware it
    would come back as a redirect to /login and Chrome would use nothing."""
    response = anon_client.get("/static/icons/manifest.json")
    assert response.status_code == 200
    assert "json" in response.headers["content-type"]

    manifest = response.json()
    assert manifest["start_url"] == "/"
    assert manifest["scope"] == "/"          # or Chrome scopes it to /static/icons
    assert manifest["display"] == "standalone"
    for icon in manifest["icons"]:
        assert anon_client.get(icon["src"]).status_code == 200


def test_no_api_docs_exposed(client):
    """docs_url=None — staff should never land on FastAPI's Swagger page."""
    assert client.get("/docs").status_code == 404


def test_unknown_route_is_not_a_crash(client):
    assert client.get("/does-not-exist").status_code == 404


def _au(iso: str) -> str:
    """Mirror of the au_date filter, so page assertions read naturally."""
    from app.main import au_date

    return au_date(iso)


def _from_au(rendered: str) -> str:
    """'14 Aug 2026' -> '2026-08-14', for asserting on order rather than text."""
    import datetime as dt

    return dt.datetime.strptime(rendered.strip(), "%d %b %Y").date().isoformat()


# ------------------------------------- a photo on the scan path, 16 Aug

def test_scanning_a_known_product_without_a_photo_offers_one(client, sample):
    """944 of the shop's 945 products had no picture, because the only place to
    attach one was behind the Edit toggle on the product screen — a trip nobody
    makes later, for an item they are no longer holding."""
    body = client.get("/scan?barcode=9300601234567").text
    assert 'type="file"' in body
    assert 'name="photo"' in body
    assert 'enctype="multipart/form-data"' in body


def test_scanning_a_new_barcode_offers_a_photo(client):
    body = client.get("/scan?barcode=9999999999999").text
    assert 'name="photo"' in body
    assert 'enctype="multipart/form-data"' in body


def test_the_scan_photo_opens_the_camera_not_a_file_browser(client, sample):
    """`capture` is the difference between "which app?" and the camera opening.

    Asserted as a contrast, because the point is that the two screens differ on
    purpose: on the product screen you may be picking a picture you already
    have, so it stays a plain chooser; on the scan path you are standing at the
    shelf holding the item, so Android should go straight to the camera.
    """
    assert 'capture="environment"' in client.get("/scan?barcode=9999999999999").text

    edit = client.get(f"/products/{sample['products']['monster']}").text
    assert 'name="photo"' in edit, "the product screen still takes a photo"
    assert "capture=" not in edit, "the product screen must stay a plain chooser"


def test_a_product_that_already_has_a_photo_is_not_asked_again(client, sample, db):
    """Replacing one stays on the product screen. The add path is used hundreds
    of times a week and does not carry a control for the rare case."""
    product = sample["products"]["monster"]
    db.execute("UPDATE products SET image_path = ? WHERE id = ?",
               ("data/photos/9300601234567.jpg", product))
    db.commit()
    body = client.get("/scan?barcode=9300601234567").text
    assert 'name="photo"' not in body


def test_a_photo_taken_while_scanning_is_saved_against_the_product(client, sample, db):
    client.post("/scan/add", data={
        "barcode": "9300609876543", "expiry_date": days(5),
    }, files={"photo": ("counter.jpg", _photo_bytes(), "image/jpeg")})
    stored = db.execute(
        "SELECT image_path FROM products WHERE barcode = ?", ("9300609876543",)
    ).fetchone()[0]
    assert stored == "data/photos/9300609876543.jpg"


def test_a_bad_photo_does_not_cost_you_the_date(client, sample, db):
    """The asymmetry that matters, and the opposite of the product screen's rule.

    There, one form means one save and a refused photo takes the name with it.
    Here the date is the entire point of the screen and the photo is a bonus, so
    the batch is written first and a picture Pillow cannot read is reported
    afterwards — never by throwing away a date that was typed correctly.
    """
    before = db.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
    response = client.post("/scan/add", data={
        "barcode": "9300609876543", "expiry_date": days(6),
    }, files={"photo": ("notes.txt", b"not a photograph", "text/plain")},
        follow_redirects=True)

    assert db.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == before + 1
    assert db.execute(
        "SELECT image_path FROM products WHERE barcode = ?", ("9300609876543",)
    ).fetchone()[0] is None
    assert "photo did not" in response.text


def test_an_enormous_image_does_not_cost_you_the_date(client, sample, db):
    """The same rule as the test above, for the failure that got past it.

    store_upload only caught OSError, which covers "this is not an image" and
    almost nothing else. Pillow refuses an image with far too many pixels by
    raising DecompressionBombError, which is not an OSError — so it came out of
    the route, the commit below it never ran, and the batch inserted moments
    earlier went with it. The person is told the app crashed, and the date they
    typed correctly is gone.

    It does not take a malicious file: a solid-colour PNG of 20000x12000 is
    under a megabyte on disk, so it sails past the size limit and only becomes
    240 million pixels once Pillow opens it.
    """
    import io

    from PIL import Image

    from app.photos import MAX_UPLOAD_BYTES

    buffer = io.BytesIO()
    Image.new("RGB", (20000, 12000), (255, 0, 0)).save(buffer, "PNG")
    bomb = buffer.getvalue()
    assert len(bomb) < MAX_UPLOAD_BYTES, "the size limit would have caught it"

    before = db.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
    response = client.post("/scan/add", data={
        "barcode": "9300609876543", "expiry_date": days(6),
    }, files={"photo": ("huge.png", bomb, "image/png")}, follow_redirects=True)

    assert response.status_code == 200
    # The date is the point of the screen. It survives a photo that does not.
    assert db.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == before + 1
    assert "photo did not" in response.text
    # And it says what actually went wrong. This was a real image, so telling
    # somebody it "was not an image" would send them looking for the wrong
    # thing — the reason travels through the redirect as a code.
    assert "too large to process" in response.text
    assert "was not an image" not in response.text


def test_photo_js_does_not_steer_the_browser_itself(client):
    """The 16 Aug Android bug, as a rule.

    photo.js used to POST the form with fetch and then move the browser to a
    destination written into the markup as `data-done`. That was right for the
    one form carrying the attribute and `undefined` for the scan form that was
    not, so a photo which saved perfectly well landed the person on
    `/undefined` and a 404. Reading `response.url` instead only moved the
    problem: on the paths where the server re-renders rather than redirects —
    "already tracked", "give it a name" — that is the POST url, and following
    it with a GET is another dead end.

    The script now swaps the shrunk file into the input and lets the browser
    submit the form, so redirects, re-rendered errors and the back button are
    all the browser's business. Asserted on the source because the failure is
    invisible from Python: every server-side test posts the form directly and
    never runs a line of this file.
    """
    # Comments stripped: this file explains the bug it used to have, and the
    # explanation names both of the things being asserted against.
    js = client.get("/static/js/photo.js").text
    code = re.sub(r"/\*.*?\*/", "", js, flags=re.DOTALL)
    code = re.sub(r"^\s*//.*$", "", code, flags=re.MULTILINE)

    assert "dataset.done" not in code, "the destination is hardcoded in the markup again"
    assert "response.url" not in code, "following the response url breaks re-rendered errors"
    assert "form.submit()" in code, "the form no longer submits itself"

    # And nothing renders the attribute any more either.
    assert "data-done" not in client.get("/scan?barcode=9999999999999").text
