"""The locked decisions from CLAUDE.md, as tests.

`scripts/check_db.py` asserts these hold in the *current* database. These
assert the database *cannot be made* to break them — the difference between
"nobody has done this yet" and "this is impossible".

If one of these fails, don't loosen it. Either the change was wrong, or the
decision genuinely moved and should be changed here deliberately, in a commit
that says so.
"""

from __future__ import annotations

import hashlib
import re
import sqlite3
from pathlib import Path

import pytest

from app.main import au_date
from app.views import au_when
from app.security import hash_pin, verify_pin
from conftest import days

ROOT = Path(__file__).resolve().parent.parent


# ------------------------------------------------------- the duplication guard

def test_duplicate_live_batch_is_refused(db, sample):
    """The whole data model. Same product, same date, both live — impossible."""
    product = sample["products"]["monster"]
    date = days(60)
    db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
               (product, date))
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
                   (product, date))


def test_discounted_still_counts_as_live(db, sample):
    """A discounted batch is still on the shelf. Don't let a twin in behind it."""
    product = sample["products"]["monster"]
    date = days(61)
    db.execute("INSERT INTO batches (product_id, expiry_date, status) VALUES (?, ?, ?)",
               (product, date, "discounted"))
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
                   (product, date))


def test_a_date_can_repeat_once_the_batch_is_deleted(db, sample):
    """Stock cleared in March can legitimately recur in September.

    Rewritten 13 Aug. This used to park a 'pulled' row on the date and prove the
    partial index let the date be reused. There is no resolved-but-present status
    any more, so the way a date comes free is deletion — and a deleted row is
    simply not there, which is why the index can stay partial and still be right.
    """
    product = sample["products"]["monster"]
    date = days(62)
    first = db.execute(
        "INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)", (product, date)
    ).lastrowid

    # Still live, so the date is taken and the guard says so.
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
                   (product, date))

    db.execute("DELETE FROM batches WHERE id = ?", (first,))
    db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
               (product, date))  # must not raise


def test_the_duplicate_check_can_exclude_the_row_being_edited(db, sample):
    """`live_batch(exclude_id=...)`, which is what makes editing a date possible.

    Tested here rather than through the screen because both layers deliberately
    produce the same message — the app catches the duplicate first, the index
    backstops it, and "nobody needs to know which layer noticed". From outside
    the route the two are indistinguishable, so the app-level check is pinned
    directly.
    """
    from app.catalogue import live_batch

    product = sample["products"]["curly"]
    edge = sample["batches"]["due_edge"]        # +7
    taken = days(1)                             # the discounted batch's date

    # Moving +7 onto +1 collides with the other batch.
    assert live_batch(db, product, taken, exclude_id=edge) is not None

    # Saving +7 back onto +7 finds only itself, which is not a collision.
    assert live_batch(db, product, days(7), exclude_id=edge) is None
    # And without the exclusion it does find itself — the bug this guards.
    assert live_batch(db, product, days(7)) is not None


def test_same_date_different_products_is_fine(db, sample):
    date = days(63)
    db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
               (sample["products"]["monster"], date))
    db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
               (sample["products"]["curly"], date))


# ------------------------------------------------------------------ categories

def test_categories_are_case_insensitively_unique(db, sample):
    """Ten people typing freely produce Drinks/drinks/DRINKS within a week."""
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO categories (name) VALUES (?)", ("ENERGY DRINKS",))


def test_a_product_may_have_no_category(db, sample):
    """NULL is a normal state, not a missing value to be filled in."""
    db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
               ("9300602222222", "Something nobody categorised"))
    db.commit()
    assert db.execute(
        "SELECT category_id FROM products WHERE barcode = '9300602222222'"
    ).fetchone()[0] is None


def test_deleting_a_category_does_not_delete_products(db, sample):
    """ON DELETE SET NULL. Losing a category must never lose the catalogue."""
    db.execute("DELETE FROM categories WHERE id = ?", (sample["cat_id"],))
    db.commit()
    row = db.execute("SELECT category_id FROM products WHERE id = ?",
                     (sample["products"]["monster"],)).fetchone()
    assert row is not None and row[0] is None


# ----------------------------------------------------------------- no counting

def test_quantity_cannot_be_zero_or_negative(db, sample):
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO batches (product_id, expiry_date, quantity) "
                   "VALUES (?, ?, ?)", (sample["products"]["monster"], days(64), 0))


def test_quantity_defaults_to_one(db, sample):
    batch_id = db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
                          (sample["products"]["monster"], days(65))).lastrowid
    assert db.execute("SELECT quantity FROM batches WHERE id = ?",
                      (batch_id,)).fetchone()[0] == 1


def test_quantity_is_never_rendered(client, sample):
    """The column exists for a possible future. It must not reach the UI."""
    assert "quantity" not in client.get("/").text.lower()


# ------------------------------------------------------------- no roles, PINs

def test_users_have_no_role_column(db):
    columns = [c[1] for c in db.execute("PRAGMA table_info(users)")]
    assert "role" not in columns


def test_pin_round_trips(db):
    pin_hash, pin_salt = hash_pin("4821")
    assert verify_pin("4821", pin_hash, pin_salt)
    assert not verify_pin("4822", pin_hash, pin_salt)


def test_pin_must_be_four_digits(db):
    for bad in ("123", "12345", "abcd", "", "12a4"):
        with pytest.raises(ValueError):
            hash_pin(bad)


def test_same_pin_gets_different_hashes(db):
    """Per-user salt. Two staff picking 1234 must not look identical in the table."""
    assert hash_pin("1234")[0] != hash_pin("1234")[0]


def test_an_account_off_the_list_cannot_add_anything(client, sample, db, staff):
    """Retiring an account has to bite the sessions already holding it.

    The cookie lasts thirty days and the middleware never looks at the
    database, so without the check in current_user an iPad still signed in as
    `BP TECOMA` would go on stamping batches with it for a month after it was
    taken off the list. That is exactly the thing this iteration exists to
    stop.
    """
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})

    def add(expiry: str) -> int:
        """Add a date to a known barcode; how many batches now carry that date."""
        client.post(
            "/scan/add",
            data={"barcode": "9300601234567", "expiry_date": expiry},
            follow_redirects=False,
        )
        return db.execute(
            "SELECT COUNT(*) FROM batches WHERE expiry_date = ?", (expiry,)
        ).fetchone()[0]

    assert add(days(45)) == 1, "the control failed — this test proves nothing"

    client.post(f"/settings/staff/{staff['id']}/active", data={"active": "0"})
    assert add(days(46)) == 0


def test_somebody_off_the_list_can_still_reach_the_sign_in_screen(client, staff):
    """Otherwise retiring your own account locks you out of your own laptop.

    /login sends a signed-in person to the home screen. Somebody just taken
    off the list is still signed in as far as the cookie goes, so without the
    active check there they would bounce straight back and never get to sign
    in as themselves.
    """
    client.post("/settings/staff", data={"name": "Sarah", "pin": "4821"})
    client.post(f"/settings/staff/{staff['id']}/active", data={"active": "0"})

    response = client.get("/login", follow_redirects=False)
    assert response.status_code == 200
    assert "Sarah" in response.text


def test_the_shop_cannot_empty_its_own_sign_in_list(client, db, staff):
    """No roles means anyone can retire anyone. Retiring the last one is a lockout."""
    body = client.post(
        f"/settings/staff/{staff['id']}/active", data={"active": "0"}, follow_redirects=True
    ).text
    assert "only person left" in body
    assert db.execute("SELECT COUNT(*) FROM users WHERE active = 1").fetchone()[0] == 1


# --------------------------------------------------------------- no CDN links

# The shop laptop has to work with the internet down. A CDN link is invisible
# until the day the connection drops, which is exactly the day someone is
# standing at the counter with a queue.

# Matches a URL only where the browser would actually go and fetch it, so
# prose in a comment ("a plain http:// address is not a secure context") does
# not trip it. That distinction matters: the rule is about requests, not text.
EXTERNAL = re.compile(
    r"""(?:src|href)\s*=\s*["']\s*(?:https?:)?//"""
    r"""|url\(\s*["']?\s*(?:https?:)?//"""
    r"""|@import\s+["']?\s*(?:https?:)?//"""
    r"""|(?:fetch|importScripts)\(\s*["']\s*(?:https?:)?//""",
    re.IGNORECASE,
)

SERVED = [
    *(ROOT / "app" / "templates").glob("*.html"),
    *(ROOT / "app" / "static" / "js").glob("*.js"),
    *(ROOT / "app" / "static" / "css").glob("*.css"),
]


def test_nothing_served_to_a_browser_links_out():
    """Every asset comes from this laptop. Vendored files go in static/vendor."""
    assert SERVED, "found no templates or assets to check — the globs are wrong"
    offenders = []
    for path in SERVED:
        for number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
            if EXTERNAL.search(line):
                offenders.append(f"{path.relative_to(ROOT)}:{number}: {line.strip()}")
    assert not offenders, "external requests found:\n" + "\n".join(offenders)


CSS = ROOT / "app" / "static" / "css" / "app.css"


def test_the_hidden_attribute_actually_hides():
    """A hidden element must not be on screen. This was not true.

    The browser's own rule is `[hidden] { display: none }` — one attribute
    selector, which any author rule setting `display` outranks. `.btn` sets
    `display: inline-block`, so the camera button shipped `hidden`, was left
    hidden on every device without a camera, and was drawn anyway: visible,
    and completely dead, because scanner.js returns before it binds a click.
    Found on an iPad over plain http, where it is the normal state.

    There is no browser in this suite, so this checks the mechanism rather
    than the pixels: the stylesheet has to re-assert `[hidden]` hard enough
    to win. Removing that rule brings the bug straight back.
    """
    # Comments stripped first: the rule is explained in a comment that quotes
    # the browser's weaker version of itself, and a plain search finds that.
    css = re.sub(r"/\*.*?\*/", "", CSS.read_text(encoding="utf-8"), flags=re.DOTALL)
    rule = re.search(r"\[hidden\]\s*\{([^}]*)\}", css)
    assert rule, "nothing in app.css re-asserts [hidden]"
    body = rule.group(1)
    assert "display" in body and "none" in body, f"[hidden] does not hide: {body!r}"
    assert "!important" in body, "without !important a .btn display rule still wins"

    # The hazard has to still be there for the guard to be worth keeping. If
    # this ever fails, .btn stopped setting display and this test can go.
    assert re.search(r"\.btn\s*\{[^}]*display\s*:", css), \
        ".btn no longer sets display — re-check whether the guard is still needed"


def test_the_vendored_scanner_is_the_file_we_vetted():
    """Pinned by checksum, so a swapped or edited bundle is not silent.

    app/static/vendor/README.md records where this came from and what it
    hashes to. Nothing in that folder is edited by hand — if this fails, the
    file changed and the question is who changed it and to what.
    """
    vendor = ROOT / "app" / "static" / "vendor"
    library = vendor / "zxing-0.21.3.min.js"
    assert library.exists(), "the vendored scanner library is missing"

    recorded = re.search(
        r"SHA-256.*?`([0-9a-f]{64})`", (vendor / "README.md").read_text(encoding="utf-8")
    )
    assert recorded, "README.md no longer records a SHA-256 for the library"

    actual = hashlib.sha256(library.read_bytes()).hexdigest()
    assert actual == recorded.group(1), (
        f"the vendored library does not match README.md\n"
        f"  recorded {recorded.group(1)}\n  actual   {actual}"
    )


# ----------------------------------------------------------------------- dates

def test_au_date_is_australian():
    assert au_date("2026-09-04") == "4 Sep 2026"


def test_au_date_has_no_leading_zero():
    """The Windows trap. strftime('%-d') does not exist there — au_date builds
    the day by hand, and this test is what stops someone 'simplifying' it."""
    assert au_date("2026-01-01") == "1 Jan 2026"
    assert not au_date("2026-01-01").startswith("01")


def test_au_date_handles_two_digit_days():
    assert au_date("2026-12-25") == "25 Dec 2026"


def test_au_date_is_never_us_format():
    """4 Sep 2026, never 9/4/2026 — the failure mode is silent and dangerous."""
    assert "/" not in au_date("2026-09-04")


def test_au_when_is_australian_and_local():
    """Stamps are stored UTC by datetime('now'); the shop reads GMT+10."""
    assert au_when("2026-09-04 03:11:06") == "4 Sep 2026, 1:11 pm"
    assert au_when("2026-01-01 14:05:00") == "2 Jan 2026, 12:05 am"


def test_au_when_has_no_leading_zero_hour():
    """The Windows trap again — '%-I' does not exist there either."""
    assert au_when("2026-09-04 22:30:00").endswith("8:30 am")
    assert ", 08:" not in au_when("2026-09-04 22:30:00")


def test_au_when_survives_an_empty_stamp():
    """resolved_at is NULL for everything nobody has resolved."""
    assert au_when(None) == ""


# ------------------------------------------------------------------- integrity

def test_foreign_keys_are_enforced(db, sample):
    """A raw sqlite3.connect() silently drops this pragma. connect() must not."""
    assert db.execute("PRAGMA foreign_keys").fetchone()[0] == 1
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO batches (product_id, expiry_date) VALUES (?, ?)",
                   (999999, days(66)))


def test_synchronous_is_full(db):
    """What makes a closed lid safe. FULL is 2."""
    assert db.execute("PRAGMA synchronous").fetchone()[0] == 2


def test_the_database_path_can_be_pointed_elsewhere(tmp_path, monkeypatch):
    """`TTD_DB`, the sibling of `TTD_PHOTO_DIR`. Unset in production.

    Added so the LAN test server can run against a copy of the shop's data. The
    actions the iPad session exercises are a real delete and a real discount, and
    without this the only way to try them on realistic data was to move
    data/tecoma.db aside by hand and hope to put it back.
    """
    import importlib

    import scripts.init_db as init_db

    # Unset, it resolves to the shop's file — the production path.
    monkeypatch.delenv("TTD_DB", raising=False)
    importlib.reload(init_db)
    assert init_db.DB_PATH == init_db.ROOT / "data" / "tecoma.db"

    # Set, everything that goes through connect() follows it.
    elsewhere = tmp_path / "copy.db"
    monkeypatch.setenv("TTD_DB", str(elsewhere))
    importlib.reload(init_db)
    assert init_db.DB_PATH == elsewhere

    conn = init_db.connect()
    conn.execute("CREATE TABLE probe (x INTEGER)")
    conn.commit()
    conn.close()
    assert elsewhere.exists(), "connect() wrote somewhere else entirely"

    # Leave the module as the rest of the suite expects to find it.
    monkeypatch.delenv("TTD_DB", raising=False)
    importlib.reload(init_db)


def test_barcodes_are_unique(db, sample):
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
                   ("9300601234567", "A different product, same barcode"))


@pytest.mark.parametrize("status", ["gone", "pulled", "sold", "Active", ""])
def test_batch_status_is_constrained(db, sample, status):
    """Two statuses, not four. 'pulled' and 'sold' are now invalid values.

    They were removed on 13 Aug because the shop never used either — 1757
    active and 583 pulled came from the import, zero sold, zero discounted — and
    because a record of every item ever removed only grows on a laptop nobody
    prunes. This is the backstop; app/routes/products.py refuses them first.
    """
    with pytest.raises(sqlite3.IntegrityError):
        db.execute("INSERT INTO batches (product_id, expiry_date, status) "
                   "VALUES (?, ?, ?)", (sample["products"]["monster"], days(67), status))


def test_a_batch_has_exactly_two_endings(db, sample):
    """Discounted, or deleted. Nothing else is representable."""
    import re

    ddl = db.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='batches'"
    ).fetchone()[0]
    # Just the CHECK clause: sqlite_master keeps the comments too, and those
    # explain what 'pulled' and 'sold' were, so searching the whole DDL finds
    # the words it is meant to prove absent.
    clause = re.search(r"CHECK\s*\(\s*status\s+IN\s*\(([^)]*)\)", ddl, re.I)
    assert clause, f"no status CHECK found in:\n{ddl}"
    values = {v.strip().strip("'\"") for v in clause.group(1).split(",")}
    assert values == {"active", "discounted"}, values


def test_products_are_not_attributed_to_a_person(db):
    """A batch is something somebody did; a product is a fact about a barcode.

    `added_by` and `resolved_by` stay on batches — they are why PINs exist.
    """
    columns = [c[1] for c in db.execute("PRAGMA table_info(products)")]
    assert "created_by" not in columns
    batch_columns = [c[1] for c in db.execute("PRAGMA table_info(batches)")]
    assert "added_by" in batch_columns and "resolved_by" in batch_columns


# The delete confirmation asks only when the item is still good AND still full
# price. Anything else, and a decision about that item has already been made.

@pytest.mark.parametrize("status,offset,expected", [
    ("active", 4, True),        # still good, full price — ask
    ("active", 1, True),
    ("active", 0, True),        # expires today, still sellable — ask
    ("active", -1, False),      # past its date — no question, it is being pulled
    ("active", -30, False),
    ("discounted", 4, False),   # already decided about, and marked down
    ("discounted", 0, False),
    ("discounted", -1, False),
])
def test_the_delete_confirmation_rule(status, offset, expected):
    from app.catalogue import needs_confirmation

    assert needs_confirmation(status, days(offset)) is expected


def test_the_confirmation_counts_whole_days(db, sample):
    from app.catalogue import days_until

    assert days_until(days(4)) == 4
    assert days_until(days(0)) == 0
    assert days_until(days(-3)) == -3
    assert days_until("not a date") is None


# ------------------------------------------------------------- the barcode rule
#
# Digits only, 6 to 18 of them, after stripping the gun's AIM identifier.
# Decided 13 Aug: a product is never deleted, so a barcode typed as a word is
# permanent. Two layers, and both are tested — app/catalogue.py produces the
# sentence staff read, the CHECK constraint makes it impossible by any route.

@pytest.mark.parametrize("value", [
    "cool ridge water",                  # the actual bug: a typed name
    "https://qrco.de/bdeRvm",            # a marketing QR, 5 of these are in the export
    "www.ausbev.com.au",
    "12345",                             # 5 digits, under the floor
    "1" * 19,                            # 19, over the ceiling
    "1930083008300830083008300830083008300830",   # the real 40-digit gun misfire
    "93006 01234567",                    # embedded space
    "9300601234567x",
    "",
    "   ",
])
def test_the_barcode_rule_refuses(value):
    from app.catalogue import parse_barcode

    barcode, problem = parse_barcode(value)
    assert barcode == "", f"{value!r} was accepted"
    assert problem, f"{value!r} was refused with no reason to show anybody"


@pytest.mark.parametrize("value,expected", [
    ("9300601234567", "9300601234567"),      # ordinary EAN-13
    ("0000001051117", "0000001051117"),      # leading zeros are significant
    ("123456", "123456"),                    # 6, the floor
    ("1" * 18, "1" * 18),                    # 18, the ceiling
    ("  9300601234567  ", "9300601234567"),  # whitespace from the gun
    ("9300601234567\r\n", "9300601234567"),  # the gun's trailing newline
    ("]C10118721274620198", "0118721274620198"),   # AIM identifier stripped
    ("]E09300601234567", "9300601234567"),
])
def test_the_barcode_rule_accepts_and_normalises(value, expected):
    """The prefix is the gun naming the symbology, not part of the code."""
    from app.catalogue import parse_barcode

    barcode, problem = parse_barcode(value)
    assert not problem, f"{value!r} was refused: {problem}"
    assert barcode == expected


def test_the_barcode_check_constraint_is_the_backstop(db, sample):
    """The app checks first; this is what makes it impossible another way."""
    for value in ("cool ridge water", "https://qrco.de/bdeRvm", "12345", "1" * 19):
        with pytest.raises(sqlite3.IntegrityError):
            db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
                       (value, "Should never exist"))


def test_the_two_barcode_layers_agree_on_unicode_digits(db, sample):
    """The subtle one. str.isdigit() is True for '٣' and '²'.

    Had the app used it, those would pass the app and then hit the CHECK
    constraint, so staff would get an IntegrityError page instead of a
    sentence. The rule has to refuse the same things in both halves.
    """
    from app.catalogue import parse_barcode

    for value in ("١٢٣٤٥٦٧٨", "9300601234567²", "½½½½½½"):
        barcode, problem = parse_barcode(value)
        assert problem, f"the app accepted {value!r}"
        # And the database would have refused it too, so they cannot disagree.
        with pytest.raises(sqlite3.IntegrityError):
            db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
                       (value, "Should never exist"))


def test_a_gun_prefixed_scan_finds_the_existing_product(client, sample):
    """Normalising on the lookup path too, or the gun invents a second product.

    `sample` already holds 9300601234567. Scanning it with the prefix attached
    must land on that product's date step, not the new-product form.
    """
    body = client.get("/scan?barcode=]C19300601234567").text
    assert "Monster Ultra Zero 500ml" in body
    assert "New barcode" not in body


def test_deleting_a_product_removes_its_batches(db, sample):
    """ON DELETE CASCADE — no orphan batches, which check_db.py also asserts."""
    db.execute("DELETE FROM products WHERE id = ?", (sample["products"]["monster"],))
    db.commit()
    assert db.execute("SELECT COUNT(*) FROM batches WHERE product_id = ?",
                      (sample["products"]["monster"],)).fetchone()[0] == 0


def test_no_categories_are_seeded(db):
    """seed.sql must stay settings-only. The list grows by itself."""
    assert db.execute("SELECT COUNT(*) FROM categories").fetchone()[0] == 0


def test_the_window_is_a_setting_not_a_constant(db):
    assert db.execute(
        "SELECT value FROM settings WHERE key = 'expiry_window_days'"
    ).fetchone()[0] == "7"


# --------------------------------------------------------------- the export
#
# A spreadsheet is the one place this data leaves the app, and Excel rewrites
# anything it is allowed to interpret. These say it may not.

def _export(conn, **kwargs):
    from io import BytesIO

    from openpyxl import load_workbook

    from scripts.export_xlsx import build

    return load_workbook(BytesIO(build(conn, **kwargs))).active


def _row_for(sheet, product_name):
    """The one row for this product. Strict on purpose.

    A product can have several batches, and returning the first match would
    let a change in sort order quietly point a test at a different row —
    which is exactly what happened once. Target products with one batch.
    """
    rows = [row for row in sheet.iter_rows(min_row=2) if row[0].value == product_name]
    assert len(rows) == 1, f"{product_name!r} has {len(rows)} rows, expected 1"
    return rows[0]


# One batch, so a row lookup is unambiguous. Trailing spaces and all.
SINGLE = "C/RIDGE WATER 1L  "


def test_a_barcode_keeps_its_leading_zeros(db, sample):
    """0000001051117 is a real product. As a number it becomes 1051117."""
    db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
               ("0000001051117", "Zero Lead Test Item"))
    db.execute("INSERT INTO batches (product_id, expiry_date) "
               "SELECT id, ? FROM products WHERE barcode = ?",
               (days(20), "0000001051117"))
    db.commit()

    cell = _row_for(_export(db), "Zero Lead Test Item")[1]
    assert cell.value == "0000001051117"
    assert isinstance(cell.value, str), "written as a number, the zeros are gone"
    assert cell.number_format == "@", "without a text format Excel re-reads it"


def test_a_long_barcode_is_not_turned_into_a_number(db, sample):
    """Rewritten 13 Aug, when the barcode rule made the old fixtures illegal.

    This used to insert ']C10118721274620198' and a QR-code URL, because both
    were in the real data. Neither can exist now — the CHECK constraint refuses
    them and the migration cleared the ones that were there.

    The risk it was guarding is not gone though, it moved: an 18-digit numeric
    barcode is exactly what Excel turns into 1.11E+17. So the case is now the
    longest barcode the rule allows, which is a stronger test than the old one
    — ']C1...' was safe from Excel precisely because it was not a number.
    """
    longest = "1" * 18
    db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
               (longest, "Eighteen Digit Item"))
    db.execute("INSERT INTO batches (product_id, expiry_date) "
               "SELECT id, ? FROM products WHERE barcode = ?", (days(21), longest))
    db.commit()

    cell = _row_for(_export(db), "Eighteen Digit Item")[1]
    assert cell.value == longest
    assert isinstance(cell.value, str), "as a number this is 1.11E+17"
    assert cell.number_format == "@", "without a text format Excel re-reads it"


def test_expiry_is_a_real_date_never_a_us_string(db, sample):
    import datetime as dt

    cell = _row_for(_export(db), SINGLE)[3]
    assert isinstance(cell.value, (dt.date, dt.datetime)), "a string is Excel's to guess at"
    assert "mmm" in cell.number_format, "a numeric month can be read either way round"
    assert "mm/dd" not in cell.number_format


def test_the_export_carries_no_quantity(db, sample):
    """There is no counting. The column exists in the schema and stays there."""
    headings = [c.value for c in _export(db)[1]]
    assert not any("quantity" in str(h).lower() for h in headings)


def test_timestamps_are_shifted_to_shop_time(db, sample):
    """SQLite stores UTC. The shop is GMT+10 and reads these as local."""
    db.execute("UPDATE batches SET added_at = ? WHERE product_id = ?",
               ("2026-08-11 03:30:00", sample["products"]["shouty"]))
    db.commit()

    stamp = _row_for(_export(db), SINGLE)[8].value
    assert (stamp.hour, stamp.minute) == (13, 30), "still UTC"
    assert stamp.day == 11


def test_a_date_only_stamp_is_not_shifted_into_a_time(db, sample):
    """A stamp that is a date, not a timestamp. +10h would invent 10am.

    The 583 migrated rows this was written for are gone with the status change,
    and the app always writes a full datetime('now'). Kept because the shift is
    still applied to whatever is in the column, and a date-only value is exactly
    the input that turns into a plausible-looking lie.
    """
    import datetime as dt

    db.execute("UPDATE batches SET status = 'discounted', resolved_at = ? WHERE id = ?",
               ("2026-03-02", sample["batches"]["due_soon"]))
    db.commit()

    cell = _row_for(_export(db), SINGLE)[10]
    assert cell.value == dt.datetime(2026, 3, 2, 0, 0), "a time appeared from nowhere"
    assert cell.number_format == "d mmm yyyy"


def test_names_go_in_exactly_as_the_shop_has_them(db, sample):
    """Messy case, trailing space, curly apostrophe. Do not clean them."""
    products = {row[0].value for row in _export(db).iter_rows(min_row=2)}
    assert "C/RIDGE WATER 1L  " in products
    assert "Arnott’s Tim Tam 200g" in products


def test_a_name_starting_with_equals_is_not_a_formula(db, sample):
    db.execute("INSERT INTO products (barcode, name) VALUES (?, ?)",
               ("9300600000999", "=SUM(A1:A9) Mints"))
    db.execute("INSERT INTO batches (product_id, expiry_date) "
               "SELECT id, ? FROM products WHERE barcode = ?",
               (days(22), "9300600000999"))
    db.commit()

    cell = _row_for(_export(db), "=SUM(A1:A9) Mints")[0]
    assert cell.data_type == "s", "Excel would show #NAME? where the product should be"


def test_exporting_changes_nothing_in_the_database(db, sample):
    """It is a copy for reading. Nothing here is a write."""
    from scripts.export_xlsx import build

    before = "\n".join(db.iterdump())
    build(db)
    assert "\n".join(db.iterdump()) == before


# ------------------------------------------------------- the uplift, 14 Aug
#
# Colour does three jobs and urgency is not one of them: green commits, red
# removes, yellow is a state and never an action. These hold that line.

TEMPLATES = sorted((ROOT / "app" / "templates").glob("*.html"))
BASE = ROOT / "app" / "templates" / "base.html"

# A control, from its opening tag to its close. <summary> counts: the Delete
# that opens a confirmation is one, and it has to be red like the rest.
CONTROL = re.compile(r"<(button|summary|a)\b([^>]*)>(.*?)</\1>", re.DOTALL)


def _controls(html: str):
    """(attributes, visible label) for every button, summary and link."""
    for _tag, attrs, inner in CONTROL.findall(html):
        yield attrs, " ".join(re.sub(r"<[^>]+>", " ", inner).split())


def test_delete_is_always_red_and_nothing_else_is(client, sample):
    """Deleting a batch is the only thing in this app that cannot be undone.

    A discount can be taken back — "Back to full price" clears it — but a
    deleted row is gone for real, which is why it is the one red in the app
    and why nothing else may borrow it. Staff had a solid red Delete in the
    old app for months; this keeps that association rather than spending it.
    """
    product = sample["products"]["curly"]
    for path in ("/", f"/products/{product}", "/settings"):
        body = client.get(path).text
        seen = 0
        for attrs, label in _controls(body):
            red = "btn-danger" in attrs
            if label in ("Delete", "Yes, delete"):
                assert red, f"{path}: {label!r} is not red"
                seen += 1
            else:
                assert not red, f"{path}: {label!r} is red and should not be"
        assert seen, f"{path}: no delete control rendered, so this proved nothing"


def test_a_destructive_control_always_says_what_it_does(client, sample):
    """An icon on its own is learned, not read.

    Swipe-to-act was built and removed on 13 Aug: a person cannot see where an
    invisible threshold is, so every swipe was a guess, and the two guesses
    were "delete" and "discount". A bare trash can is the same mistake in a
    different shape. Every red control keeps its word beside the glyph.
    """
    product = sample["products"]["curly"]
    for path in ("/", f"/products/{product}", "/settings"):
        for attrs, label in _controls(client.get(path).text):
            if "btn-danger" in attrs:
                assert re.search(r"[A-Za-z]", label), \
                    f"{path}: a red control with no word on it — icon only"


def test_every_icon_points_at_a_symbol_that_exists(client, sample):
    """A typo in an icon name is silent: <use> just draws nothing.

    So the button keeps its word and its box and loses only the glyph, which
    is exactly the kind of thing nobody notices until a photo of the shop
    screen turns up. Both directions: no reference without a symbol, and no
    symbol nobody uses, because the sprite ships on every page.
    """
    defined = set(re.findall(r'<symbol id="(i-[a-z-]+)"', BASE.read_text(encoding="utf-8")))
    assert defined, "no icons defined in base.html — the sprite went missing"

    used = set()
    for path in TEMPLATES:
        used |= set(re.findall(r'href="#(i-[a-z-]+)"', path.read_text(encoding="utf-8")))

    assert used - defined == set(), f"icons used with no symbol: {sorted(used - defined)}"
    assert defined - used == set(), f"icons defined and never used: {sorted(defined - used)}"


def test_an_icon_never_carries_an_aria_role(client, sample):
    """`role` is a forbidden word on the settings screen, and the sprite is on
    every page — so an ARIA role on an icon would fail a test three screens
    away from the edit. The icons are decorative and say so with aria-hidden;
    the word beside them is what a screen reader announces.
    """
    for path in TEMPLATES:
        body = path.read_text(encoding="utf-8")
        assert not re.search(r"\brole\s*=", body), f"{path.name} carries a role attribute"


def test_the_home_screen_icon_opens_in_safari_for_the_camera():
    """iOS blocks the live scanning camera in a standalone home-screen web app
    on the shop's iPads, so the icon has to open in Safari instead —
    `apple-mobile-web-app-capable` is "no", not "yes".

    Flipping it back to "yes" looks like a nicer full-screen app and silently
    kills the aisle scan on every iPad: getUserMedia is absent there, so the
    "Scan with camera" button never even appears. The Android scanner phone has
    no such restriction and keeps display:standalone via the manifest — this is
    an iOS-only knob.
    """
    base = (ROOT / "app" / "templates" / "base.html").read_text(encoding="utf-8")
    assert 'name="apple-mobile-web-app-capable"' in base
    assert re.search(
        r'name="apple-mobile-web-app-capable"\s+content="no"', base
    ), "the home-screen icon is standalone again — the iPad camera will not open"


def test_yellow_marks_a_state_and_is_never_an_action():
    """The discount sticker is the one saturated yellow in the app, and it
    works because it is the only one. The moment a button takes that colour
    the sticker stops being a state you can spot from down the aisle and
    starts being decoration.
    """
    css = re.sub(r"/\*.*?\*/", "", CSS.read_text(encoding="utf-8"), flags=re.DOTALL)
    for block in re.findall(r"([^{}]+)\{([^}]*)\}", css):
        selector, body = block[0].strip(), block[1]
        if ".btn" in selector and "--pump" in body:
            assert False, f"a button is wearing the sticker's yellow: {selector}"
    assert "--pump" in css, "the sticker colour is gone entirely"


def test_the_printed_sheet_still_costs_no_toner():
    """A solid yellow sticker and a page of icons are free on glass and a
    toner bill on paper. The sheet is printed every week.
    """
    css = CSS.read_text(encoding="utf-8")
    block = css.split("@media print")[1]
    assert ".tag-discount" in block, "the sticker prints as a solid fill"
    assert re.search(r"\.tag,\s*\.tag-discount\s*\{[^}]*background:\s*none", block), \
        "the sticker's background is not cleared for print"
    assert re.search(r"\.icon\s*\{[^}]*display:\s*none", block), "icons print"


def test_keyboard_focus_is_visible():
    """The laptop has no touchscreen and gets driven by tab and enter during a
    stocktake. Before the uplift nothing in the stylesheet styled focus at
    all, so there was nothing to follow.

    Anchored to the bare `:focus-visible`, not merely to some rule containing
    it — the bar has its own `.bar :focus-visible` that swaps the colour, and
    matching that one would let the app-wide rule be deleted unnoticed.
    """
    css = re.sub(r"/\*.*?\*/", "", CSS.read_text(encoding="utf-8"), flags=re.DOTALL)
    rule = re.search(r"(?:^|\})\s*:focus-visible\s*\{([^}]*)\}", css)
    assert rule, "nothing in app.css gives every focusable thing a focus ring"
    assert re.search(r"outline\s*:", rule.group(1)), \
        f"focus is styled but not outlined: {rule.group(1)!r}"


# ------------------------------------------------- the day bands, 16 Aug

def test_the_day_bands_read_as_words_not_numbers():
    """"Today" and "Tomorrow", not "0 days left" and "1 day left".

    The delete confirmation already says "this expires today" / "tomorrow", and
    two parts of the same screen must not describe the same day two ways.
    """
    from app.catalogue import days_left_label

    assert days_left_label(0) == "Today"
    assert days_left_label(1) == "Tomorrow"
    assert days_left_label(2) == "2 days left"
    assert days_left_label(7) == "7 days left"
    # Past-date rows are their own group and cannot reach here, but a negative
    # must never render as "-2 days left" if one ever does.
    assert days_left_label(-2) == "Today"


def test_the_later_heading_follows_the_window_setting():
    """"More than a week" is only true at a 7-day window.

    The window is a setting, so a shop running 10 days would otherwise be told
    that 9 days away is "more than a week".
    """
    from app.catalogue import later_heading

    assert later_heading(7) == "More than a week"
    assert later_heading(3) == "More than 3 days"
    assert later_heading(14) == "More than 14 days"


def test_the_name_outweighs_the_date_in_a_row():
    """Reversed 16 Aug, so it is a rule now rather than an accident.

    The band above the row states the urgency, so the name — the thing being
    matched against an item in a hand — carries the weight, and the date sits
    back. Asserted on the declarations themselves: a later edit that quietly
    swaps them back should go red here rather than in somebody's eyes.
    """
    css = re.sub(r"/\*.*?\*/", "", CSS.read_text(encoding="utf-8"), flags=re.DOTALL)

    def weight(selector: str) -> int:
        rule = re.search(r"(?:^|\})\s*\.%s\s*\{([^}]*)\}" % selector, css)
        assert rule, f".{selector} has no rule in app.css"
        found = re.search(r"font-weight:\s*(\d+)", rule.group(1))
        assert found, f".{selector} does not set a font-weight"
        return int(found.group(1))

    assert weight("item-name") > weight("item-date"), \
        "the date is back to outweighing the product name"


def test_the_bar_stays_put_without_position_sticky():
    """The shop's handheld is Chrome 50, which sits in the gap where Chrome had
    no `position: sticky` at all (removed in 37, restored in 56). Without a
    fallback the bar scrolls away there, which on the one device carried round
    the aisles means the navigation leaves the screen.

    Asserted three ways, because each is a different way to break it: the
    fallback must exist, it must not be reachable by print (a bounded, scrolling
    `main` would put page one on the paper and drop the rest), and it must not
    hardcode the bar's height — that number is 62px on the laptop and 85px on a
    phone, and a wrong one hides content behind the bar forever.
    """
    css = re.sub(r"/\*.*?\*/", "", CSS.read_text(encoding="utf-8"), flags=re.DOTALL)

    start = css.find("@supports not (position: sticky)")
    assert start != -1, "nothing keeps the bar put where sticky is unsupported"

    depth, end = 0, None
    for i in range(css.index("{", start), len(css)):
        if css[i] == "{":
            depth += 1
        elif css[i] == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    assert end, "the @supports block is not closed"
    block = css[start:end]

    assert "@media screen" in block, \
        "the fallback is not scoped to screen, so printing the sheet can clip it"
    assert re.search(r"main\s*\{[^}]*overflow-y:\s*auto", block), \
        "the fallback does not give main its own scroller"
    assert not re.search(r"padding-top:\s*\d", block), \
        "the fallback hardcodes a bar height, which is not one number"
