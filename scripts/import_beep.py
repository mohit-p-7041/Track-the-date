"""Import an Excel export from the old beep app into the Track the Date database.

    python scripts/import_beep.py data/imports/beep_2026-08-10.xlsx
    python scripts/import_beep.py <file> --dry-run     # report only, write nothing

Expected columns: Id, Name, Barcode, Expiration Date, Category, Memo,
Added User, Added Date.

What it does with the data:

  * One product per unique barcode. Names are trimmed; the first name seen for
    a barcode wins and any later disagreement is reported.
  * One batch per (barcode, expiry date) pair. The shop does not count stock —
    a second item is only recorded if its date differs — so rows that repeat a
    pair collapse into one batch rather than accumulating a quantity.
  * Products import with no category. The old app had only one category ('All'),
    so there is nothing to migrate. Staff assign categories as they rescan
    products during normal work.

  * Rows already past their expiry date are imported as status 'pulled', with a
    note recording that this was the migration and not a person. The stock is
    long gone from the shelves; the records just never got cleared.
  * Staff named in 'Added User' are created as users so the audit trail carries
    over. They all get the same placeholder PIN — change it before go-live.
    There are no roles; everyone can do everything.
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.catalogue import parse_barcode  # noqa: E402
from app.security import hash_pin  # noqa: E402
from scripts.init_db import DB_PATH, connect  # noqa: E402

PLACEHOLDER_PIN = "1234"
MIGRATION_NOTE = "Expired before migration — not verified"

EXPECTED_HEADERS = ["Id", "Name", "Barcode", "Expiration Date", "Category", "Memo",
                    "Added User", "Added Date"]


def _clean_barcode(value) -> str:
    """Barcodes arrive as floats or strings; normalise to digits, keep leading zeros."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip()


def _clean_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(str(value).strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def read_rows(xlsx: Path) -> list[dict]:
    try:
        import warnings
        with warnings.catch_warnings():
            # the beep export carries no default style; openpyxl warns and copes fine
            warnings.simplefilter("ignore")
            import openpyxl
    except ImportError:
        sys.exit("openpyxl is not installed. Run: pip install -r requirements.txt")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"{xlsx} is empty.")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        sys.exit(f"Export is missing expected columns: {missing}\nFound: {headers}")

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS}
    out = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append({h: r[i] for h, i in idx.items()})
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--db", type=Path, default=DB_PATH)
    ap.add_argument("--today", type=str, default=None,
                    help="override today's date (YYYY-MM-DD), for testing")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"No such file: {args.xlsx}")
    if not args.db.exists():
        sys.exit(f"No database at {args.db}. Run: python scripts/init_db.py")

    today = (dt.date.fromisoformat(args.today) if args.today else dt.date.today())
    rows = read_rows(args.xlsx)

    # ---------------------------------------------------------- pass 1: shape
    products: dict[str, str] = {}           # barcode -> name
    name_conflicts: dict[str, set] = defaultdict(set)
    pairs: dict[tuple, dict] = {}           # (barcode, date) -> batch info
    staff: set[str] = set()
    skipped: list[tuple[int, str]] = []
    refused_barcodes: dict[str, int] = defaultdict(int)   # raw barcode -> rows

    for n, row in enumerate(rows, start=2):
        barcode = _clean_barcode(row["Barcode"])
        name = (str(row["Name"]).strip() if row["Name"] else "")
        expiry = _clean_date(row["Expiration Date"])

        if not barcode:
            skipped.append((n, "no barcode"))
            continue

        # The barcode rule, from app/catalogue.py so there is one copy of it.
        # Added 13 Aug with the CHECK constraint: without this the importer
        # dies on the export's ten non-numeric barcodes, and the laptop deploy
        # runs the importer rather than copying the database.
        #
        # This normalises before it judges, so the gun's ']C1' rows are
        # recovered under their real digits rather than dropped.
        barcode, refused = parse_barcode(barcode)
        if refused:
            refused_barcodes[_clean_barcode(row["Barcode"])] += 1
            skipped.append((n, f"barcode refused — {refused}"))
            continue

        if not expiry:
            skipped.append((n, "unreadable expiry date"))
            continue
        if not name:
            name = f"Unknown product {barcode}"

        if barcode in products:
            if products[barcode].casefold() != name.casefold():
                name_conflicts[barcode].update({products[barcode], name})
        else:
            products[barcode] = name

        user = (str(row["Added User"]).strip() if row["Added User"] else "")
        if user:
            staff.add(user)

        memo = (str(row["Memo"]).strip() if row["Memo"] else "") or None
        key = (barcode, expiry)
        if key in pairs:
            # The shop doesn't count stock: a repeated (product, date) pair is
            # a duplicate entry in the old app, not two tracked items.
            pairs[key]["dupes"] += 1
        else:
            pairs[key] = {"dupes": 0, "note": memo, "user": user,
                          "added_at": row["Added Date"]}

    merged = sum(p["dupes"] for p in pairs.values())
    expired = sum(1 for (_, d) in pairs if dt.date.fromisoformat(d) < today)

    # ------------------------------------------------------------- report
    print(f"Read {len(rows)} rows from {args.xlsx.name}\n")
    print(f"  unique products (barcodes) : {len(products)}")
    print(f"  batches to create          : {len(pairs)}")
    print(f"    already expired -> pulled: {expired}")
    print(f"    still live               : {len(pairs) - expired}")
    print(f"  duplicate rows collapsed   : {merged}")
    print(f"  staff found                : {sorted(staff)}")
    print(f"  rows skipped               : {len(skipped)}")
    for n, why in skipped[:10]:
        print(f"      row {n}: {why}")
    if refused_barcodes:
        rows_lost = sum(refused_barcodes.values())
        print(f"  barcodes refused by the rule: {len(refused_barcodes)}"
              f" ({rows_lost} rows) — digits only, 6 to 18")
        for raw, count in sorted(refused_barcodes.items(), key=lambda kv: -kv[1]):
            print(f"      {raw[:40]!r}: {count} row(s)")
    if name_conflicts:
        print(f"  barcodes with >1 name      : {len(name_conflicts)} (first seen wins)")
        for bc, names in list(name_conflicts.items())[:10]:
            print(f"      {bc}: {sorted(names)}")

    if args.dry_run:
        print("\nDry run — nothing written.")
        return 0

    # ------------------------------------------------------------- write
    conn = connect(args.db)
    try:
        user_ids: dict[str, int] = {}
        for person in sorted(staff):
            existing = conn.execute("SELECT id FROM users WHERE name = ?", (person,)).fetchone()
            if existing:
                user_ids[person] = existing[0]
                continue
            pin_hash, salt = hash_pin(PLACEHOLDER_PIN)
            cur = conn.execute(
                "INSERT INTO users (name, pin_hash, pin_salt) VALUES (?, ?, ?)",
                (person, pin_hash, salt),
            )
            user_ids[person] = cur.lastrowid

        product_ids: dict[str, int] = {}
        for barcode, name in products.items():
            existing = conn.execute(
                "SELECT id FROM products WHERE barcode = ?", (barcode,)
            ).fetchone()
            if existing:
                product_ids[barcode] = existing[0]
                continue
            # category_id stays NULL — staff categorise as they rescan.
            cur = conn.execute(
                "INSERT INTO products (barcode, name) VALUES (?, ?)", (barcode, name)
            )
            product_ids[barcode] = cur.lastrowid

        created = 0
        already = 0
        for (barcode, expiry), info in pairs.items():
            status = "pulled" if dt.date.fromisoformat(expiry) < today else "active"

            # Can't lean on INSERT OR IGNORE here: the unique index is partial
            # (live rows only) by design, so re-importing would duplicate every
            # already-expired batch. Check explicitly, whatever the status.
            exists = conn.execute(
                "SELECT 1 FROM batches WHERE product_id = ? AND expiry_date = ?",
                (product_ids[barcode], expiry),
            ).fetchone()
            if exists:
                already += 1
                continue

            added_at = info["added_at"]
            added_at = (added_at.isoformat(sep=" ", timespec="seconds")
                        if isinstance(added_at, dt.datetime) else None)

            # resolved_by is deliberately left NULL on pre-expired rows: nobody
            # confirmed these, so don't put a name against them.
            note = info["note"]
            if status == "pulled":
                note = f"{note}. {MIGRATION_NOTE}" if note else MIGRATION_NOTE

            conn.execute(
                """INSERT INTO batches
                   (product_id, expiry_date, note, status, added_by, added_at, resolved_at)
                   VALUES (?, ?, ?, ?, ?, COALESCE(?, datetime('now')), ?)""",
                (product_ids[barcode], expiry, note, status,
                 user_ids.get(info["user"]), added_at,
                 today.isoformat() if status == "pulled" else None),
            )
            created += 1

        conn.commit()

        live = conn.execute(
            "SELECT COUNT(*) FROM batches WHERE status IN ('active','discounted')"
        ).fetchone()[0]
        print(f"\nWrote {len(product_ids)} products and {created} batches.")
        if already:
            print(f"  {already} batches already existed and were left alone")
        print(f"  live batches now on the home screen: {live}")
        print(f"\nAll staff PINs are set to {PLACEHOLDER_PIN} — change these before go-live.")
        print("No categories were created. Staff assign them as they rescan products.")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
