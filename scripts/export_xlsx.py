"""Export every recorded date to a single Excel sheet.

    python scripts/export_xlsx.py                  # data/exports/tecoma-YYYY-MM-DD.xlsx
    python scripts/export_xlsx.py somewhere.xlsx   # or a path of your choosing

Also behind the Export button on the settings screen, which hands the same
bytes straight to the browser as a download rather than leaving a file on the
laptop. One implementation either way.

This is for the manager, not for safety — the startup backup is what protects
the data (SPEC §9). Nothing here writes to the database.

Every batch is included. Amended 13 Aug: there is no longer a resolved-but-
present row to include, because a batch now ends either as `discounted` — still
on the shelf, still live — or as a real deletion. That makes this file the only
place the shop's history is kept, so take it before either migration and keep
it. Filter the Status column to separate discounted stock from full price.

Four things about the file matter more than they look:

  * The barcode is written as text. 952 products include codes like
    0000001051117, and Excel turns a 13-digit number into 9.31E+12 and eats
    the leading zero. As text it survives, and so do the scanner's odd ones —
    ]C10118721274620198 and a handful of QR URLs are real rows in this data.
  * Expiry is a real date, not a string, so Excel sorts and filters it. Given
    a string it would guess, and its guess on this machine is US order.
  * Timestamps are shifted to GMT+10 first. SQLite stores UTC and nobody in
    Tecoma thinks in UTC. The same shift app/views.py does for the screens.
  * Product names go in exactly as stored — messy case, odd punctuation and
    all. Staff recognise them as they are; see CLAUDE.md.
"""

from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.views import SHOP_OFFSET, au_date  # noqa: E402
from scripts.init_db import DB_PATH, connect  # noqa: E402

EXPORT_DIR = ROOT / "data" / "exports"

MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

DATE_FORMAT = "d mmm yyyy"                    # 14 Aug 2026 — never 8/14/2026
STAMP_FORMAT = "d mmm yyyy h:mm AM/PM"
TEXT_FORMAT = "@"

# Heading, column width, and whether the column is forced to text. The widths
# are deliberate: the longest product name in the shop's data is 123
# characters and a column that wide is unreadable, so it is capped and Excel
# clips the tail rather than pushing everything else off the screen.
COLUMNS = [
    ("Product", 42, False),
    ("Barcode", 21, True),
    ("Category", 16, False),
    ("Expiry", 13, False),
    ("Days left", 10, False),
    ("Status", 12, False),
    ("Note", 34, False),
    ("Added by", 16, False),
    ("Added at", 20, False),
    ("Resolved by", 16, False),
    ("Resolved at", 14, False),
]

QUERY = """
    SELECT p.name        AS product,
           p.barcode     AS barcode,
           c.name        AS category,
           b.expiry_date AS expiry,
           b.status      AS status,
           b.note        AS note,
           ua.name       AS added_by,
           b.added_at    AS added_at,
           ur.name       AS resolved_by,
           b.resolved_at AS resolved_at
      FROM batches b
      JOIN products p    ON p.id  = b.product_id
 LEFT JOIN categories c  ON c.id  = p.category_id
 LEFT JOIN users ua      ON ua.id = b.added_by
 LEFT JOIN users ur      ON ur.id = b.resolved_by
  ORDER BY b.expiry_date,
           p.name COLLATE NOCASE
"""
# Soonest first, the same as every screen. The old sort lifted live rows above
# resolved ones, because 583 already-pulled rows were the oldest in the file and
# put the manager's cursor on the part that no longer mattered. Those rows are
# gone with the status change, so plain date order is now the useful order.


def _as_date(value: str | None) -> dt.date | str | None:
    """A real date where the text is one, otherwise leave it alone.

    Every expiry in the database is ISO — the add path refuses anything else
    and the importer normalised the rest. Falling back to the raw string means
    a row that somehow isn't still appears in the export instead of stopping
    it, which is the right way round for a file somebody is reading to find a
    problem.
    """
    if not value:
        return None
    text = str(value)
    try:
        return dt.date.fromisoformat(text)
    except ValueError:
        return text


def _as_stamp(value: str | None) -> dt.datetime | dt.date | str | None:
    """A timestamp shifted to shop time, or a plain date left as a date.

    `added_at` is SQLite's `datetime('now')` — UTC, to the second. But the 583
    rows that came across from the old app carry a date-only `resolved_at`,
    and shifting one of those by ten hours would invent a 10am that nobody
    recorded. Length decides which it is.
    """
    if not value:
        return None
    text = str(value)
    if len(text) == 10:
        return _as_date(text)
    try:
        return dt.datetime.fromisoformat(text.replace("Z", "")) + SHOP_OFFSET
    except ValueError:
        return text


def _write(cell, value, text_format: bool = False) -> None:
    """Put a value in a cell without letting Excel reinterpret it.

    A string starting with `=` becomes a formula on assignment, and the
    manager would open the file to `#NAME?` where a product should be. No name
    in the shop's data starts that way, but staff type these names on the scan
    screen and this costs two lines.
    """
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    if text_format:
        cell.number_format = TEXT_FORMAT
    elif isinstance(value, dt.datetime):
        cell.number_format = STAMP_FORMAT
    elif isinstance(value, dt.date):
        cell.number_format = DATE_FORMAT


def build(conn: sqlite3.Connection, today: dt.date | None = None) -> bytes:
    """The workbook, as bytes. Reads the database and changes nothing."""
    today = today or dt.date.today()
    rows = conn.execute(QUERY).fetchall()

    wb = Workbook()
    ws = wb.active
    # The tab carries the day it was taken, so "Days left" can be read months
    # later without wondering what it counted from. Sheet names are capped at
    # 31 characters; this is 18.
    ws.title = f"Expiry {au_date(today)}"

    header = Font(bold=True)
    for index, (heading, width, _text) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=index, value=heading)
        cell.font = header
        cell.alignment = Alignment(vertical="bottom")
        ws.column_dimensions[get_column_letter(index)].width = width

    for line, row in enumerate(rows, start=2):
        expiry = _as_date(row["expiry"])
        days_left = (expiry - today).days if isinstance(expiry, dt.date) else None
        values = [
            row["product"],
            row["barcode"],
            row["category"],
            expiry,
            days_left,
            row["status"],
            row["note"],
            row["added_by"],
            _as_stamp(row["added_at"]),
            row["resolved_by"],
            _as_stamp(row["resolved_at"]),
        ]
        for index, (value, (_h, _w, text_format)) in enumerate(zip(values, COLUMNS), start=1):
            _write(ws.cell(row=line, column=index), value, text_format)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.properties.creator = "Track the Date — Tecoma"
    wb.properties.title = f"Expiry dates as at {au_date(today)}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def filename(today: dt.date | None = None) -> str:
    return f"tecoma-{(today or dt.date.today()).isoformat()}.xlsx"


def main() -> int:
    ap = argparse.ArgumentParser(description="Export every recorded date to Excel.")
    ap.add_argument("out", nargs="?", type=Path, help="where to write it")
    args = ap.parse_args()

    if not DB_PATH.exists():
        print(f"No database at {DB_PATH}. Run scripts/init_db.py first.")
        return 1

    out = args.out
    if out is None:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        out = EXPORT_DIR / filename()
    elif out.is_dir():
        out = out / filename()

    conn = connect()
    try:
        data = build(conn)
        rows = conn.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
    finally:
        conn.close()

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(data)
    print(f"Wrote {out} ({len(data) / 1024:.0f} KB, {rows} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
